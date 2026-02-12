"""
COMPETITOR_INTEL Tab - Complete 7-Day Competitor Intelligence
Added as TAB 8 to v3.6.0 Enhanced Dashboard

Sections in this tab:
1. 7-DAY POSTING LOG - Every competitor post, what day/time, momentum, trend
2. POSTING PATTERNS - Day-of-week and hour-of-day frequency analysis
3. RESPONSE TIME - How fast they jump on spiking trends vs you  
4. TREND SELECTION ACCURACY - What % of their picks hit high momentum
5. NICHE COVERAGE - Which categories they cover vs you
6. TEMPLATE VARIATION STRATEGY - Volume per trend analysis
7. REVENUE ESTIMATION - Estimated earnings per trend
8. CROSS-MARKET TIMING - Which market they prioritize for BOTH trends
9. WIN/LOSS SCORECARD - Weekly you vs them trend-by-trend
"""

import pandas as pd
import json
import os
from datetime import datetime, timedelta
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

# Import shared config
try:
    from v35_enhancements import YOUR_ACCOUNTS, COMPETITOR_ACCOUNTS, _sanitize_cell
except ImportError:
    from v35_enhancements_v360 import YOUR_ACCOUNTS, COMPETITOR_ACCOUNTS, _sanitize_cell


# =============================================================================
# 7-DAY COMPETITOR HISTORY CACHE
# =============================================================================

def save_competitor_history(df_today, cache_dir):
    """
    Save today's competitor posts to a rolling 7-day cache.
    Each day stored as separate file: competitor_history_YYYY-MM-DD.json
    Old files (>7 days) are automatically cleaned up.
    """
    os.makedirs(cache_dir, exist_ok=True)
    date_str = datetime.now().strftime('%Y-%m-%d')
    
    comp_lower = [a.lower() for a in COMPETITOR_ACCOUNTS]
    your_lower = [a.lower() for a in YOUR_ACCOUNTS]
    
    # Extract competitor posts with full detail
    if 'author' not in df_today.columns:
        print("  [CompIntel] No author column — skipping competitor cache save")
        return
    
    comp_mask = df_today['author'].str.lower().isin(comp_lower)
    comp_posts = df_today[comp_mask].copy()
    
    # Also save YOUR posts for comparison
    your_mask = df_today['author'].str.lower().isin(your_lower)
    your_posts = df_today[your_mask].copy()
    
    def _to_serializable(df):
        """Convert DataFrame to JSON-safe list of dicts."""
        records = []
        for _, row in df.iterrows():
            record = {}
            for col in ['webVideoUrl', 'author', 'text', 'momentum_score', 
                        'shares_per_hour', 'views_per_hour', 'likes_per_hour',
                        'age_hours', 'Market', 'AI_CATEGORY', 'acceleration_status',
                        'shareCount', 'playCount', 'diggCount', 'createTimeISO']:
                val = row.get(col, None)
                if pd.isna(val):
                    record[col] = None
                elif isinstance(val, (int, float)):
                    record[col] = float(val) if isinstance(val, float) else int(val)
                else:
                    record[col] = str(val)
            records.append(record)
        return records
    
    cache_data = {
        'date': date_str,
        'competitor_posts': _to_serializable(comp_posts),
        'your_posts': _to_serializable(your_posts),
    }
    
    cache_path = os.path.join(cache_dir, f'competitor_history_{date_str}.json')
    with open(cache_path, 'w') as f:
        json.dump(cache_data, f)
    print(f"  [CompIntel] Saved {len(comp_posts)} competitor + {len(your_posts)} your posts to {cache_path}")
    
    # Cleanup files older than 7 days
    cutoff = datetime.now() - timedelta(days=8)
    for fn in os.listdir(cache_dir):
        if fn.startswith('competitor_history_') and fn.endswith('.json'):
            try:
                file_date = datetime.strptime(fn.replace('competitor_history_', '').replace('.json', ''), '%Y-%m-%d')
                if file_date < cutoff:
                    os.remove(os.path.join(cache_dir, fn))
                    print(f"  [CompIntel] Cleaned up old cache: {fn}")
            except ValueError:
                pass


def load_competitor_history(cache_dir, days=7):
    """
    Load up to `days` days of competitor history.
    Returns dict: { 'YYYY-MM-DD': { 'competitor_posts': [...], 'your_posts': [...] }, ... }
    """
    history = {}
    if not os.path.exists(cache_dir):
        return history
    
    for fn in sorted(os.listdir(cache_dir)):
        if fn.startswith('competitor_history_') and fn.endswith('.json'):
            try:
                file_date = fn.replace('competitor_history_', '').replace('.json', '')
                # Only load last N days
                d = datetime.strptime(file_date, '%Y-%m-%d')
                if d >= datetime.now() - timedelta(days=days):
                    with open(os.path.join(cache_dir, fn), 'r') as f:
                        history[file_date] = json.load(f)
            except (ValueError, json.JSONDecodeError) as e:
                print(f"  [CompIntel] Skipping corrupt cache file {fn}: {e}")
    
    print(f"  [CompIntel] Loaded {len(history)} days of competitor history")
    return history


# =============================================================================
# ANALYSIS FUNCTIONS
# =============================================================================

def build_competitor_intel(df_today, cache_dir):
    """
    Build all competitor intel data structures from today's data + 7-day history.
    Returns dict with all analysis sections ready for Excel rendering.
    """
    history = load_competitor_history(cache_dir)
    
    comp_lower = [a.lower() for a in COMPETITOR_ACCOUNTS]
    your_lower = [a.lower() for a in YOUR_ACCOUNTS]
    
    # --- Combine all historical competitor posts ---
    all_comp_posts = []  # list of dicts with 'date' added
    all_your_posts = []
    
    for date_str, day_data in history.items():
        for post in day_data.get('competitor_posts', []):
            post['cache_date'] = date_str
            all_comp_posts.append(post)
        for post in day_data.get('your_posts', []):
            post['cache_date'] = date_str
            all_your_posts.append(post)
    
    # Also include today's live data (may not be cached yet)
    today_str = datetime.now().strftime('%Y-%m-%d')
    if today_str not in history and 'author' in df_today.columns:
        comp_mask = df_today['author'].str.lower().isin(comp_lower)
        for _, row in df_today[comp_mask].iterrows():
            all_comp_posts.append({
                'cache_date': today_str,
                'webVideoUrl': str(row.get('webVideoUrl', '')),
                'author': str(row.get('author', '')),
                'text': str(row.get('text', ''))[:80] if pd.notna(row.get('text')) else '',
                'momentum_score': float(row.get('momentum_score', 0)),
                'shares_per_hour': float(row.get('shares_per_hour', 0)),
                'views_per_hour': float(row.get('views_per_hour', 0)),
                'likes_per_hour': float(row.get('likes_per_hour', 0)),
                'age_hours': float(row.get('age_hours', 0)),
                'Market': str(row.get('Market', '')),
                'AI_CATEGORY': str(row.get('AI_CATEGORY', '')),
                'acceleration_status': str(row.get('acceleration_status', '')),
                'createTimeISO': str(row.get('createTimeISO', '')),
            })
        your_mask = df_today['author'].str.lower().isin(your_lower)
        for _, row in df_today[your_mask].iterrows():
            all_your_posts.append({
                'cache_date': today_str,
                'webVideoUrl': str(row.get('webVideoUrl', '')),
                'author': str(row.get('author', '')),
                'text': str(row.get('text', ''))[:80] if pd.notna(row.get('text')) else '',
                'momentum_score': float(row.get('momentum_score', 0)),
                'shares_per_hour': float(row.get('shares_per_hour', 0)),
                'views_per_hour': float(row.get('views_per_hour', 0)),
                'age_hours': float(row.get('age_hours', 0)),
                'Market': str(row.get('Market', '')),
                'AI_CATEGORY': str(row.get('AI_CATEGORY', '')),
                'createTimeISO': str(row.get('createTimeISO', '')),
            })
    
    comp_df = pd.DataFrame(all_comp_posts) if all_comp_posts else pd.DataFrame()
    your_df = pd.DataFrame(all_your_posts) if all_your_posts else pd.DataFrame()
    
    intel = {}
    
    # --- SECTION 1: 7-DAY POSTING LOG ---
    intel['posting_log'] = _analyze_posting_log(comp_df)
    
    # --- SECTION 2: POSTING PATTERNS ---
    intel['posting_patterns'] = _analyze_posting_patterns(comp_df)
    
    # --- SECTION 3: RESPONSE TIME TO SPIKING ---
    intel['response_time'] = _analyze_response_time(comp_df, your_df)
    
    # --- SECTION 4: TREND SELECTION ACCURACY ---
    intel['selection_accuracy'] = _analyze_selection_accuracy(comp_df, your_df)
    
    # --- SECTION 5: NICHE COVERAGE ---
    intel['niche_coverage'] = _analyze_niche_coverage(comp_df, your_df)
    
    # --- SECTION 6: TEMPLATE VARIATION STRATEGY ---
    intel['variation_strategy'] = _analyze_variation_strategy(comp_df, your_df)
    
    # --- SECTION 7: REVENUE ESTIMATION ---
    intel['revenue_estimate'] = _analyze_revenue_estimation(comp_df)
    
    # --- SECTION 8: CROSS-MARKET TIMING ---
    intel['cross_market'] = _analyze_cross_market(comp_df, your_df)
    
    # --- SECTION 9: WIN/LOSS SCORECARD ---
    intel['win_loss'] = _analyze_win_loss(comp_df, your_df, history)
    
    intel['days_of_data'] = len(history)
    
    return intel


def _analyze_posting_log(comp_df):
    """Section 1: Every competitor post, day, time, momentum."""
    if len(comp_df) == 0:
        return []
    
    rows = []
    for _, p in comp_df.iterrows():
        # Extract post time from createTimeISO
        post_time = ''
        post_day = ''
        create_iso = p.get('createTimeISO', '')
        if create_iso and create_iso != 'None':
            try:
                dt = pd.to_datetime(create_iso)
                post_day = dt.strftime('%A')  # Monday, Tuesday, etc.
                post_time = dt.strftime('%H:%M')  # 14:30
            except Exception:
                pass
        
        rows.append({
            'date_seen': p.get('cache_date', ''),
            'account': p.get('author', ''),
            'trend': str(p.get('text', ''))[:60],
            'post_day': post_day,
            'post_time': post_time,
            'momentum': float(p.get('momentum_score', 0)),
            'shares_h': float(p.get('shares_per_hour', 0)),
            'views_h': float(p.get('views_per_hour', 0)),
            'age_h': float(p.get('age_hours', 0)),
            'market': p.get('Market', ''),
            'ai_cat': p.get('AI_CATEGORY', ''),
            'status': p.get('acceleration_status', ''),
            'url': p.get('webVideoUrl', '')
        })
    
    # Sort by date seen (newest first), then momentum (highest first)
    rows.sort(key=lambda x: (-_date_rank(x['date_seen']), -x['momentum']))
    return rows


def _date_rank(date_str):
    """Convert date string to sortable number."""
    try:
        return int(date_str.replace('-', ''))
    except Exception:
        return 0


def _analyze_posting_patterns(comp_df):
    """Section 2: Day-of-week and hour-of-day frequency analysis."""
    if len(comp_df) == 0:
        return {'by_day': {}, 'by_hour': {}, 'busiest_day': 'N/A', 'busiest_hour': 'N/A', 'posts_per_day_avg': 0}
    
    day_counts = {}
    hour_counts = {}
    account_day_counts = {}  # per-account day breakdown
    
    for _, p in comp_df.iterrows():
        create_iso = p.get('createTimeISO', '')
        if create_iso and create_iso != 'None':
            try:
                dt = pd.to_datetime(create_iso)
                day_name = dt.strftime('%A')
                hour = dt.hour
                day_counts[day_name] = day_counts.get(day_name, 0) + 1
                hour_counts[hour] = hour_counts.get(hour, 0) + 1
                
                acct = p.get('author', 'unknown')
                if acct not in account_day_counts:
                    account_day_counts[acct] = {}
                account_day_counts[acct][day_name] = account_day_counts[acct].get(day_name, 0) + 1
            except Exception:
                pass
    
    busiest_day = max(day_counts, key=day_counts.get) if day_counts else 'N/A'
    busiest_hour = max(hour_counts, key=hour_counts.get) if hour_counts else 'N/A'
    
    # Calculate avg posts per day
    unique_dates = comp_df['cache_date'].nunique() if 'cache_date' in comp_df.columns else 1
    avg_per_day = len(comp_df) / max(unique_dates, 1)
    
    return {
        'by_day': day_counts,
        'by_hour': hour_counts,
        'by_account': account_day_counts,
        'busiest_day': busiest_day,
        'busiest_hour': f"{busiest_hour}:00 UTC" if isinstance(busiest_hour, int) else busiest_hour,
        'posts_per_day_avg': round(avg_per_day, 1),
        'total_posts': len(comp_df),
        'unique_dates': unique_dates
    }


def _analyze_response_time(comp_df, your_df):
    """Section 3: How fast competitors jump on spiking trends vs you."""
    if len(comp_df) == 0:
        return {'comp_avg_age': 'N/A', 'your_avg_age': 'N/A', 'speed_advantage': 'N/A', 'details': []}
    
    # Average age when posting (lower = faster to jump on trends)
    comp_ages = [float(p.get('age_hours', 0)) for _, p in comp_df.iterrows() if float(p.get('age_hours', 0)) > 0]
    your_ages = [float(p.get('age_hours', 0)) for _, p in your_df.iterrows() if float(p.get('age_hours', 0)) > 0]
    
    comp_avg = sum(comp_ages) / len(comp_ages) if comp_ages else 0
    your_avg = sum(your_ages) / len(your_ages) if your_ages else 0
    
    # Who's faster?
    if comp_avg > 0 and your_avg > 0:
        diff = your_avg - comp_avg
        if diff > 0:
            speed_note = f"They're {abs(diff):.1f}h faster on average"
        elif diff < 0:
            speed_note = f"You're {abs(diff):.1f}h faster on average"
        else:
            speed_note = "Tied"
    else:
        speed_note = 'Insufficient data'
    
    # Per-account breakdown
    acct_speeds = {}
    for _, p in comp_df.iterrows():
        acct = p.get('author', '')
        age = float(p.get('age_hours', 0))
        if age > 0:
            if acct not in acct_speeds:
                acct_speeds[acct] = []
            acct_speeds[acct].append(age)
    
    acct_avgs = {acct: round(sum(ages)/len(ages), 1) for acct, ages in acct_speeds.items()}
    
    return {
        'comp_avg_age': round(comp_avg, 1),
        'your_avg_age': round(your_avg, 1),
        'speed_advantage': speed_note,
        'comp_by_account': acct_avgs,
    }


def _analyze_selection_accuracy(comp_df, your_df):
    """Section 4: What % of their trend picks hit high momentum."""
    if len(comp_df) == 0:
        return {'comp_hit_rate': 'N/A', 'your_hit_rate': 'N/A'}
    
    # "Hit" = momentum >= 1000 (WATCH threshold)
    # "Big hit" = momentum >= 2000 (HIGH threshold)
    # "Massive" = momentum >= 3000 (URGENT threshold)
    
    def _calc_rates(df, label):
        total = len(df)
        if total == 0:
            return {'total': 0, 'hits_1000': 0, 'hits_2000': 0, 'hits_3000': 0, 
                    'hit_rate': 0, 'big_hit_rate': 0, 'massive_rate': 0, 'avg_momentum': 0}
        
        moms = [float(p.get('momentum_score', 0)) for _, p in df.iterrows()]
        hits = sum(1 for m in moms if m >= 1000)
        big = sum(1 for m in moms if m >= 2000)
        massive = sum(1 for m in moms if m >= 3000)
        
        return {
            'total': total,
            'hits_1000': hits,
            'hits_2000': big,
            'hits_3000': massive,
            'hit_rate': round(hits / total * 100, 1),
            'big_hit_rate': round(big / total * 100, 1),
            'massive_rate': round(massive / total * 100, 1),
            'avg_momentum': round(sum(moms) / total, 0)
        }
    
    return {
        'competitor': _calc_rates(comp_df, 'comp'),
        'you': _calc_rates(your_df, 'you'),
    }


def _analyze_niche_coverage(comp_df, your_df):
    """Section 5: AI vs NON-AI category split comparison."""
    if len(comp_df) == 0:
        return {'comp_split': {}, 'your_split': {}, 'gaps': []}
    
    def _split(df):
        if len(df) == 0:
            return {}
        cats = {}
        for _, p in df.iterrows():
            cat = p.get('AI_CATEGORY', 'Unknown')
            cats[cat] = cats.get(cat, 0) + 1
        total = sum(cats.values())
        return {k: {'count': v, 'pct': round(v/total*100, 1)} for k, v in cats.items()}
    
    comp_split = _split(comp_df)
    your_split = _split(your_df)
    
    # Identify gaps — categories they cover more heavily than you
    gaps = []
    for cat in comp_split:
        comp_pct = comp_split[cat]['pct']
        your_pct = your_split.get(cat, {}).get('pct', 0)
        if comp_pct > your_pct + 10:  # They cover 10%+ more of this category
            gaps.append(f"Competitor heavier on {cat}: {comp_pct}% vs your {your_pct}%")
    
    # Market coverage comparison
    def _market_split(df):
        if len(df) == 0:
            return {}
        markets = {}
        for _, p in df.iterrows():
            m = str(p.get('Market', 'Unknown'))
            markets[m] = markets.get(m, 0) + 1
        total = sum(markets.values())
        return {k: {'count': v, 'pct': round(v/total*100, 1)} for k, v in markets.items()}
    
    return {
        'comp_ai_split': comp_split,
        'your_ai_split': your_split,
        'comp_market_split': _market_split(comp_df),
        'your_market_split': _market_split(your_df),
        'gaps': gaps
    }


def _analyze_variation_strategy(comp_df, your_df):
    """Section 6: How many templates/posts per unique trend."""
    # Group by similar trend text (rough: first 30 chars)
    def _count_variations(df):
        if len(df) == 0:
            return {'avg_per_trend': 0, 'max_per_trend': 0, 'total_unique': 0, 'details': []}
        
        # Group by URL (same video = same trend)
        url_counts = {}
        for _, p in df.iterrows():
            url = p.get('webVideoUrl', '')
            if url not in url_counts:
                url_counts[url] = {'count': 0, 'accounts': set(), 'text': str(p.get('text', ''))[:50]}
            url_counts[url]['count'] += 1
            url_counts[url]['accounts'].add(p.get('author', ''))
        
        counts = [v['count'] for v in url_counts.values()]
        multi = [v for v in url_counts.values() if v['count'] > 1]
        
        return {
            'avg_per_trend': round(sum(counts) / len(counts), 1) if counts else 0,
            'max_per_trend': max(counts) if counts else 0,
            'total_unique': len(url_counts),
            'multi_account_trends': len(multi),
            'details': sorted([{'text': v['text'], 'count': v['count'], 'accounts': len(v['accounts'])} 
                              for v in url_counts.values()], key=lambda x: -x['count'])[:10]
        }
    
    return {
        'competitor': _count_variations(comp_df),
        'you': _count_variations(your_df),
    }


def _analyze_revenue_estimation(comp_df):
    """Section 7: Estimated competitor earnings per trend."""
    if len(comp_df) == 0:
        return {'total_estimated': 0, 'per_trend': []}
    
    # Rough model: £5 per 1000 momentum points, capped at $2500 per template
    per_trend = []
    for _, p in comp_df.iterrows():
        mom = float(p.get('momentum_score', 0))
        est = min((mom / 1000) * 5, 2500)
        per_trend.append({
            'account': p.get('author', ''),
            'trend': str(p.get('text', ''))[:50],
            'momentum': mom,
            'est_revenue': round(est, 2),
            'market': p.get('Market', ''),
            'date': p.get('cache_date', ''),
        })
    
    per_trend.sort(key=lambda x: -x['est_revenue'])
    total = sum(t['est_revenue'] for t in per_trend)
    
    return {
        'total_estimated_7d': round(total, 2),
        'per_trend': per_trend[:20],  # Top 20 by estimated revenue
        'top_earner': per_trend[0] if per_trend else None,
    }


def _analyze_cross_market(comp_df, your_df):
    """Section 8: Which market competitors prioritize for BOTH trends."""
    if len(comp_df) == 0:
        return {'comp_both_count': 0, 'your_both_count': 0}
    
    def _both_stats(df, label):
        both = [p for _, p in df.iterrows() if 'BOTH' in str(p.get('Market', ''))]
        us_only = [p for _, p in df.iterrows() if 'US' in str(p.get('Market', '')) and 'BOTH' not in str(p.get('Market', ''))]
        uk_only = [p for _, p in df.iterrows() if 'UK' in str(p.get('Market', '')) and 'BOTH' not in str(p.get('Market', ''))]
        return {
            'both': len(both),
            'us_only': len(us_only),
            'uk_only': len(uk_only),
            'both_pct': round(len(both) / max(len(df), 1) * 100, 1)
        }
    
    return {
        'competitor': _both_stats(comp_df, 'comp'),
        'you': _both_stats(your_df, 'you'),
    }


def _analyze_win_loss(comp_df, your_df, history):
    """Section 9: Win/loss scorecard - who caught more unique high-momentum trends."""
    if len(comp_df) == 0:
        return {'wins': 0, 'losses': 0, 'draws': 0, 'neither': 0, 'details': []}
    
    # Get all unique trend URLs across both
    comp_urls = set(p.get('webVideoUrl', '') for _, p in comp_df.iterrows())
    your_urls = set(p.get('webVideoUrl', '') for _, p in your_df.iterrows())
    
    all_urls = comp_urls | your_urls
    
    wins = 0    # You caught it, they didn't
    losses = 0  # They caught it, you didn't
    draws = 0   # Both caught it
    details = []
    
    # Build momentum lookup
    comp_mom = {}
    for _, p in comp_df.iterrows():
        url = p.get('webVideoUrl', '')
        mom = float(p.get('momentum_score', 0))
        if url not in comp_mom or mom > comp_mom[url]:
            comp_mom[url] = mom
    
    your_mom = {}
    for _, p in your_df.iterrows():
        url = p.get('webVideoUrl', '')
        mom = float(p.get('momentum_score', 0))
        if url not in your_mom or mom > your_mom[url]:
            your_mom[url] = mom
    
    for url in all_urls:
        in_comp = url in comp_urls
        in_yours = url in your_urls
        
        mom = max(comp_mom.get(url, 0), your_mom.get(url, 0))
        
        if in_comp and in_yours:
            draws += 1
            result = 'DRAW'
        elif in_yours and not in_comp:
            wins += 1
            result = 'WIN'
        else:
            losses += 1
            result = 'LOSS'
        
        if mom >= 500:  # Only track meaningful trends
            details.append({
                'url': url,
                'result': result,
                'momentum': mom,
                'your_momentum': your_mom.get(url, 0),
                'comp_momentum': comp_mom.get(url, 0),
            })
    
    details.sort(key=lambda x: -x['momentum'])
    
    return {
        'wins': wins,
        'losses': losses,
        'draws': draws,
        'total': wins + losses + draws,
        'win_rate': round(wins / max(wins + losses, 1) * 100, 1),
        'details': details[:20]
    }


# =============================================================================
# EXCEL TAB BUILDER
# =============================================================================

def build_competitor_intel_tab(ws, intel, header_fill, header_font, thin_border):
    """
    Build the COMPETITOR_INTEL tab with 9 sections.
    """
    row = 1
    
    # Color palette
    section_fill = PatternFill('solid', fgColor='1F4E78')
    section_font = Font(bold=True, color='FFFFFF', size=12)
    subsection_fill = PatternFill('solid', fgColor='E8F4FD')
    subsection_font = Font(bold=True, size=11)
    stat_fill = PatternFill('solid', fgColor='F5F5F5')
    win_fill = PatternFill('solid', fgColor='90EE90')
    loss_fill = PatternFill('solid', fgColor='FF6B6B')
    draw_fill = PatternFill('solid', fgColor='FFFDE0')
    gold_fill = PatternFill('solid', fgColor='FFD700')
    orange_fill = PatternFill('solid', fgColor='FFE4B5')
    cyan_fill = PatternFill('solid', fgColor='E0FFFF')
    green_fill = PatternFill('solid', fgColor='E0FFE0')
    red_fill_light = PatternFill('solid', fgColor='FFE0E0')
    alt_fill = PatternFill('solid', fgColor='F9F9F9')
    
    days_available = intel.get('days_of_data', 0)
    
    # ===== TITLE ROW =====
    c = ws.cell(row=row, column=1, value=f'COMPETITOR INTELLIGENCE — {days_available}-Day Analysis')
    c.fill = section_fill; c.font = Font(bold=True, color='FFFFFF', size=14)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
    row += 1
    c = ws.cell(row=row, column=1, value=f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")} | Tracking: {", ".join(COMPETITOR_ACCOUNTS)}')
    c.font = Font(italic=True, color='666666', size=9)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
    row += 2
    
    # ===== SECTION 1: 7-DAY POSTING LOG =====
    row = _write_section_header(ws, row, '📋 SECTION 1: 7-DAY COMPETITOR POSTING LOG', section_fill, section_font)
    
    log_headers = ['Date Seen', 'Account', 'Trend', 'Post Day', 'Post Time', 
                   'Momentum', 'Shares/h', 'Views/h', 'Age (h)', 'Market', 'AI/NON-AI', 'Status', 'URL']
    row = _write_headers(ws, row, log_headers, header_fill, header_font)
    
    posting_log = intel.get('posting_log', [])
    if posting_log:
        for i, entry in enumerate(posting_log[:50]):  # Cap at 50 rows
            vals = [entry.get('date_seen',''), entry.get('account',''), entry.get('trend',''),
                    entry.get('post_day',''), entry.get('post_time',''),
                    int(entry.get('momentum',0)), round(entry.get('shares_h',0), 1),
                    int(entry.get('views_h',0)), round(entry.get('age_h',0), 1),
                    entry.get('market',''), entry.get('ai_cat',''), entry.get('status',''),
                    entry.get('url','')]
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=_sanitize_cell(val))
                c.border = thin_border
                if i % 2 == 1:
                    c.fill = alt_fill
            # Color momentum cell based on value
            mom_cell = ws.cell(row=row, column=6)
            mom = int(entry.get('momentum', 0))
            if mom >= 3000:
                mom_cell.fill = PatternFill('solid', fgColor='FF0000')
                mom_cell.font = Font(bold=True, color='FFFFFF')
            elif mom >= 2000:
                mom_cell.fill = orange_fill
                mom_cell.font = Font(bold=True)
            elif mom >= 1000:
                mom_cell.fill = draw_fill
            # Hyperlink URL
            url_cell = ws.cell(row=row, column=13)
            url_val = entry.get('url', '')
            if url_val and str(url_val).startswith('http'):
                url_cell.hyperlink = str(url_val)
                url_cell.font = Font(color='0000FF', underline='single')
            row += 1
    else:
        ws.cell(row=row, column=1, value='No competitor posts found. Run for 2+ days to populate.').font = Font(italic=True, color='999999')
        row += 1
    
    row += 1  # spacer
    
    # ===== SECTION 2: POSTING PATTERNS =====
    row = _write_section_header(ws, row, '📊 SECTION 2: POSTING PATTERNS', section_fill, section_font)
    
    patterns = intel.get('posting_patterns', {})
    
    # Summary stats
    stats = [
        ('Total Posts (7d)', patterns.get('total_posts', 0)),
        ('Days of Data', patterns.get('unique_dates', 0)),
        ('Avg Posts/Day', patterns.get('posts_per_day_avg', 0)),
        ('Busiest Day', patterns.get('busiest_day', 'N/A')),
        ('Busiest Hour', patterns.get('busiest_hour', 'N/A')),
    ]
    for label, val in stats:
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = Font(bold=True); c1.fill = subsection_fill
        c2 = ws.cell(row=row, column=2, value=_sanitize_cell(val))
        c2.fill = stat_fill; c2.font = Font(bold=True)
        row += 1
    row += 1
    
    # Day-of-week breakdown
    by_day = patterns.get('by_day', {})
    if by_day:
        row = _write_subsection(ws, row, 'Posts by Day of Week', subsection_fill, subsection_font)
        for day_name in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']:
            count = by_day.get(day_name, 0)
            c1 = ws.cell(row=row, column=1, value=day_name)
            c2 = ws.cell(row=row, column=2, value=count)
            c3 = ws.cell(row=row, column=3, value='█' * min(count, 30))
            c3.font = Font(color='1F4E78')
            if count == max(by_day.values()):
                c1.fill = gold_fill; c2.fill = gold_fill
            row += 1
    row += 1
    
    # Hour-of-day breakdown
    by_hour = patterns.get('by_hour', {})
    if by_hour:
        row = _write_subsection(ws, row, 'Posts by Hour (UTC)', subsection_fill, subsection_font)
        for hour in range(24):
            count = by_hour.get(hour, 0)
            if count > 0:
                c1 = ws.cell(row=row, column=1, value=f'{hour:02d}:00')
                c2 = ws.cell(row=row, column=2, value=count)
                c3 = ws.cell(row=row, column=3, value='█' * min(count, 30))
                c3.font = Font(color='1F4E78')
                if count == max(by_hour.values()):
                    c1.fill = gold_fill; c2.fill = gold_fill
                row += 1
    row += 1
    
    # ===== SECTION 3: RESPONSE TIME =====
    row = _write_section_header(ws, row, '⏱️ SECTION 3: RESPONSE TIME TO TRENDS', section_fill, section_font)
    
    resp = intel.get('response_time', {})
    stats = [
        ('Competitor Avg Age at Post', f"{resp.get('comp_avg_age', 'N/A')}h"),
        ('Your Avg Age at Post', f"{resp.get('your_avg_age', 'N/A')}h"),
        ('Speed Verdict', resp.get('speed_advantage', 'N/A')),
    ]
    for label, val in stats:
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = Font(bold=True); c1.fill = subsection_fill
        c2 = ws.cell(row=row, column=2, value=_sanitize_cell(str(val)))
        c2.fill = stat_fill
        if 'faster' in str(val).lower() and 'you' in str(val).lower():
            c2.fill = green_fill; c2.font = Font(bold=True, color='006600')
        elif 'faster' in str(val).lower() and 'they' in str(val).lower():
            c2.fill = red_fill_light; c2.font = Font(bold=True, color='CC0000')
        row += 1
    
    # Per-account response times
    acct_speeds = resp.get('comp_by_account', {})
    if acct_speeds:
        row += 1
        row = _write_subsection(ws, row, 'Response Time by Competitor Account', subsection_fill, subsection_font)
        for acct, avg_age in sorted(acct_speeds.items(), key=lambda x: x[1]):
            c1 = ws.cell(row=row, column=1, value=acct)
            c2 = ws.cell(row=row, column=2, value=f'{avg_age}h avg')
            if avg_age < 24:
                c2.fill = green_fill
            elif avg_age < 48:
                c2.fill = draw_fill
            else:
                c2.fill = red_fill_light
            row += 1
    row += 1
    
    # ===== SECTION 4: TREND SELECTION ACCURACY =====
    row = _write_section_header(ws, row, '🎯 SECTION 4: TREND SELECTION ACCURACY', section_fill, section_font)
    
    acc = intel.get('selection_accuracy', {})
    comp_acc = acc.get('competitor', {})
    your_acc = acc.get('you', {})
    
    acc_headers = ['Metric', 'Competitor', 'You', 'Verdict']
    row = _write_headers(ws, row, acc_headers, header_fill, header_font)
    
    metrics = [
        ('Total Posts', comp_acc.get('total', 0), your_acc.get('total', 0)),
        ('Hits (≥1000 mom)', f"{comp_acc.get('hits_1000',0)} ({comp_acc.get('hit_rate',0)}%)", 
                              f"{your_acc.get('hits_1000',0)} ({your_acc.get('hit_rate',0)}%)"),
        ('Big Hits (≥2000)', f"{comp_acc.get('hits_2000',0)} ({comp_acc.get('big_hit_rate',0)}%)", 
                              f"{your_acc.get('hits_2000',0)} ({your_acc.get('big_hit_rate',0)}%)"),
        ('Massive (≥3000)', f"{comp_acc.get('hits_3000',0)} ({comp_acc.get('massive_rate',0)}%)", 
                             f"{your_acc.get('hits_3000',0)} ({your_acc.get('massive_rate',0)}%)"),
        ('Avg Momentum', int(comp_acc.get('avg_momentum',0)), int(your_acc.get('avg_momentum',0))),
    ]
    for metric_name, comp_val, your_val in metrics:
        # Determine verdict
        try:
            c_num = float(str(comp_val).split('(')[0].strip().replace('%',''))
            y_num = float(str(your_val).split('(')[0].strip().replace('%',''))
            verdict = '✅ You' if y_num > c_num else ('⚠️ Them' if c_num > y_num else '🟰 Tied')
        except Exception:
            verdict = ''
        
        vals = [metric_name, comp_val, your_val, verdict]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=_sanitize_cell(val))
            c.border = thin_border
        row += 1
    row += 1
    
    # ===== SECTION 5: NICHE COVERAGE =====
    row = _write_section_header(ws, row, '📂 SECTION 5: NICHE COVERAGE', section_fill, section_font)
    
    niche = intel.get('niche_coverage', {})
    
    # AI/NON-AI split
    row = _write_subsection(ws, row, 'AI vs NON-AI Split', subsection_fill, subsection_font)
    niche_headers = ['Category', 'Competitor', 'You']
    row = _write_headers(ws, row, niche_headers, header_fill, header_font)
    
    comp_ai = niche.get('comp_ai_split', {})
    your_ai = niche.get('your_ai_split', {})
    for cat in set(list(comp_ai.keys()) + list(your_ai.keys())):
        comp_info = comp_ai.get(cat, {})
        your_info = your_ai.get(cat, {})
        vals = [cat, 
                f"{comp_info.get('count',0)} ({comp_info.get('pct',0)}%)",
                f"{your_info.get('count',0)} ({your_info.get('pct',0)}%)"]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=_sanitize_cell(val))
            c.border = thin_border
        row += 1
    
    # Market split
    row += 1
    row = _write_subsection(ws, row, 'Market Coverage', subsection_fill, subsection_font)
    niche_headers = ['Market', 'Competitor', 'You']
    row = _write_headers(ws, row, niche_headers, header_fill, header_font)
    
    comp_mkt = niche.get('comp_market_split', {})
    your_mkt = niche.get('your_market_split', {})
    for mkt in set(list(comp_mkt.keys()) + list(your_mkt.keys())):
        comp_info = comp_mkt.get(mkt, {})
        your_info = your_mkt.get(mkt, {})
        vals = [mkt,
                f"{comp_info.get('count',0)} ({comp_info.get('pct',0)}%)",
                f"{your_info.get('count',0)} ({your_info.get('pct',0)}%)"]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=_sanitize_cell(val))
            c.border = thin_border
        row += 1
    
    # Gaps
    gaps = niche.get('gaps', [])
    if gaps:
        row += 1
        row = _write_subsection(ws, row, '⚠️ Coverage Gaps (Where they outweigh you)', subsection_fill, subsection_font)
        for gap in gaps:
            c = ws.cell(row=row, column=1, value=_sanitize_cell(gap))
            c.fill = red_fill_light
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            row += 1
    row += 1
    
    # ===== SECTION 6: TEMPLATE VARIATION STRATEGY =====
    row = _write_section_header(ws, row, '🔄 SECTION 6: TEMPLATE VARIATION STRATEGY', section_fill, section_font)
    
    var = intel.get('variation_strategy', {})
    comp_var = var.get('competitor', {})
    your_var = var.get('you', {})
    
    stats = [
        ('Unique Trends Covered', comp_var.get('total_unique', 0), your_var.get('total_unique', 0)),
        ('Avg Posts per Trend', comp_var.get('avg_per_trend', 0), your_var.get('avg_per_trend', 0)),
        ('Max Posts on Single Trend', comp_var.get('max_per_trend', 0), your_var.get('max_per_trend', 0)),
        ('Multi-Account Trends', comp_var.get('multi_account_trends', 0), your_var.get('multi_account_trends', 0)),
    ]
    var_headers = ['Metric', 'Competitor', 'You']
    row = _write_headers(ws, row, var_headers, header_fill, header_font)
    for metric_name, comp_val, your_val in stats:
        vals = [metric_name, comp_val, your_val]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=_sanitize_cell(val))
            c.border = thin_border
        row += 1
    row += 1
    
    # ===== SECTION 7: REVENUE ESTIMATION =====
    row = _write_section_header(ws, row, '💰 SECTION 7: ESTIMATED COMPETITOR REVENUE (7-Day)', section_fill, section_font)
    
    rev = intel.get('revenue_estimate', {})
    
    total_est = rev.get('total_estimated_7d', 0)
    c1 = ws.cell(row=row, column=1, value='Estimated 7-Day Revenue')
    c1.font = Font(bold=True, size=12); c1.fill = subsection_fill
    c2 = ws.cell(row=row, column=2, value=f'${total_est:,.0f}')
    c2.font = Font(bold=True, size=12, color='006600'); c2.fill = green_fill
    row += 1
    
    top_earner = rev.get('top_earner')
    if top_earner:
        c1 = ws.cell(row=row, column=1, value='Top Earning Trend')
        c1.font = Font(bold=True); c1.fill = subsection_fill
        c2 = ws.cell(row=row, column=2, value=f"${top_earner['est_revenue']:,.0f} — {top_earner['trend']}")
        c2.fill = gold_fill
        row += 1
    row += 1
    
    per_trend = rev.get('per_trend', [])
    if per_trend:
        rev_headers = ['Account', 'Trend', 'Momentum', 'Est Revenue ($)', 'Market', 'Date']
        row = _write_headers(ws, row, rev_headers, header_fill, header_font)
        for i, t in enumerate(per_trend[:15]):
            vals = [t.get('account',''), t.get('trend',''), int(t.get('momentum',0)),
                    round(t.get('est_revenue',0), 2), t.get('market',''), t.get('date','')]
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=_sanitize_cell(val))
                c.border = thin_border
                if i % 2 == 1:
                    c.fill = alt_fill
            row += 1
    row += 1
    
    # ===== SECTION 8: CROSS-MARKET TIMING =====
    row = _write_section_header(ws, row, '🌍 SECTION 8: CROSS-MARKET ANALYSIS', section_fill, section_font)
    
    cross = intel.get('cross_market', {})
    comp_cross = cross.get('competitor', {})
    your_cross = cross.get('you', {})
    
    cross_headers = ['Market', 'Competitor Posts', 'Your Posts']
    row = _write_headers(ws, row, cross_headers, header_fill, header_font)
    
    for label, comp_key, your_key in [('🌐 BOTH', 'both', 'both'), ('🇺🇸 US Only', 'us_only', 'us_only'), ('🇬🇧 UK Only', 'uk_only', 'uk_only')]:
        vals = [label, comp_cross.get(comp_key, 0), your_cross.get(your_key, 0)]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=_sanitize_cell(val))
            c.border = thin_border
        if label == '🌐 BOTH':
            ws.cell(row=row, column=1).fill = gold_fill
        row += 1
    
    # BOTH percentage comparison
    comp_both_pct = comp_cross.get('both_pct', 0)
    your_both_pct = your_cross.get('both_pct', 0)
    row += 1
    c1 = ws.cell(row=row, column=1, value='BOTH Market Focus')
    c1.font = Font(bold=True); c1.fill = subsection_fill
    c2 = ws.cell(row=row, column=2, value=f"Comp: {comp_both_pct}%")
    c3 = ws.cell(row=row, column=3, value=f"You: {your_both_pct}%")
    if your_both_pct > comp_both_pct:
        c3.fill = green_fill
    elif comp_both_pct > your_both_pct:
        c2.fill = green_fill
    row += 2
    
    # ===== SECTION 9: WIN/LOSS SCORECARD =====
    row = _write_section_header(ws, row, '🏆 SECTION 9: WIN/LOSS SCORECARD (7-Day)', section_fill, section_font)
    
    wl = intel.get('win_loss', {})
    
    # Summary cards
    wins = wl.get('wins', 0)
    losses = wl.get('losses', 0)
    draws = wl.get('draws', 0)
    
    score_labels = [
        ('✅ WINS (You caught, they missed)', wins, win_fill),
        ('❌ LOSSES (They caught, you missed)', losses, loss_fill),
        ('🟰 DRAWS (Both caught)', draws, draw_fill),
        ('Win Rate', f"{wl.get('win_rate', 0)}%", gold_fill),
    ]
    for label, val, fill in score_labels:
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = Font(bold=True, size=11); c1.fill = subsection_fill
        c2 = ws.cell(row=row, column=2, value=_sanitize_cell(val))
        c2.font = Font(bold=True, size=14); c2.fill = fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1)
        row += 1
    row += 1
    
    # Detailed win/loss table (high-momentum trends only)
    details = wl.get('details', [])
    if details:
        wl_headers = ['Result', 'Your Momentum', 'Comp Momentum', 'URL']
        row = _write_headers(ws, row, wl_headers, header_fill, header_font)
        for d in details[:15]:
            result = d.get('result', '')
            vals = [result, int(d.get('your_momentum', 0)), int(d.get('comp_momentum', 0)), d.get('url', '')]
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=_sanitize_cell(val))
                c.border = thin_border
            result_cell = ws.cell(row=row, column=1)
            if result == 'WIN':
                result_cell.fill = win_fill; result_cell.font = Font(bold=True)
            elif result == 'LOSS':
                result_cell.fill = loss_fill; result_cell.font = Font(bold=True, color='FFFFFF')
            elif result == 'DRAW':
                result_cell.fill = draw_fill
            # Hyperlink
            url_cell = ws.cell(row=row, column=4)
            url_val = d.get('url', '')
            if url_val and str(url_val).startswith('http'):
                url_cell.hyperlink = str(url_val)
                url_cell.font = Font(color='0000FF', underline='single')
            row += 1
    
    # Column widths
    widths = {'A': 35, 'B': 25, 'C': 50, 'D': 15, 'E': 15, 'F': 15,
              'G': 15, 'H': 15, 'I': 12, 'J': 15, 'K': 12, 'L': 12, 'M': 50}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    
    ws.freeze_panes = 'A2'


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def _write_section_header(ws, row, title, fill, font):
    """Write a full-width section header."""
    c = ws.cell(row=row, column=1, value=_sanitize_cell(title))
    c.fill = fill; c.font = font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
    return row + 1


def _write_subsection(ws, row, title, fill, font):
    """Write a subsection header."""
    c = ws.cell(row=row, column=1, value=_sanitize_cell(title))
    c.fill = fill; c.font = font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    return row + 1


def _write_headers(ws, row, headers, fill, font):
    """Write header row."""
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = fill; c.font = font; c.alignment = Alignment(horizontal='center')
    return row + 1
