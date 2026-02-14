"""
TikTok Trend System v3.5.0 Enhancements
- Velocity Prediction: Predict where trends are heading in 6h, 12h, 24h
- Competitor Analysis: Identify gaps and opportunities vs capcutdailyuk

Integrates with existing v3.4.0 daily_processor.py
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
import json
import os
import re as _re
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


def _sanitize_cell(value):
    """Sanitize a value before writing to an Excel cell.
    Removes illegal XML characters that openpyxl rejects."""
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub('', value)
    return value

# =============================================================================
# CONFIGURATION
# =============================================================================

YOUR_ACCOUNTS = [
    'capcuttemplates833', 'capcuttrends02', 'capcuttemplatesai',
    'artemiscc_capcut', 'capcutaistudio', 'artemiscccapcut', 'capcut.vorlagen101'
]

COMPETITOR_ACCOUNTS = [
    'capcutdailyuk', 'capcut__creations', 'jyoung101capcut',
    'capcut_templatetrends', 'capcut_core', 'capcut.trends.uk1'
]

# Velocity thresholds for predictions
VELOCITY_THRESHOLDS = {
    'EXPLOSIVE': 200,    # momentum increasing >200/day - will peak within 24h
    'STRONG': 100,       # momentum increasing >100/day - strong growth
    'MODERATE': 50,      # momentum increasing >50/day - healthy growth  
    'WEAK': 0,           # momentum flat or declining slightly
    'DECLINING': -50,    # momentum dropping
    'CRASHING': -100     # rapid decline
}

# Status colors (from v3.3.0 spec)
STATUS_COLORS = {
    '🆕 NEW': 'FFFFE0',
    '🚀 SPIKING': '00FF00',
    '📈 RISING': '90EE90',
    '📉 COOLING': 'FFB6C1',
    '❄️ DYING': 'FF0000'
}


# Historical revenue data - 59 entries from TIKTOK_SYSTEM_DASHBAORD.xlsx (2026-02-12)
# Used as seed when no dashboard file exists in automated environment
SEED_REVENUE_DATA = [
    {'TikTok URL': 'https://www.tiktok.com/@7597546182721670422/video/7597546182721670422', 'Account': '', 'Received ($)': 2500, 'Estimated ($)': 2500, 'US & EU3 Installs': 2059, 'ROW Installs': 8418, 'Engagements': 0, 'Total Installs': 10477, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7598141227619388694/video/7598141227619388694', 'Account': '', 'Received ($)': 2500, 'Estimated ($)': 2500, 'US & EU3 Installs': 839, 'ROW Installs': 2046, 'Engagements': 0, 'Total Installs': 2885, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7596436774478548246/video/7596436774478548246', 'Account': '', 'Received ($)': 2500, 'Estimated ($)': 2500, 'US & EU3 Installs': 540, 'ROW Installs': 861, 'Engagements': 0, 'Total Installs': 1401, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7597736583386762518/video/7597736583386762518', 'Account': '', 'Received ($)': 2500, 'Estimated ($)': 2500, 'US & EU3 Installs': 837, 'ROW Installs': 2202, 'Engagements': 0, 'Total Installs': 3039, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7599034433294830870/video/7599034433294830870', 'Account': '', 'Received ($)': 2500, 'Estimated ($)': 2500, 'US & EU3 Installs': 1399, 'ROW Installs': 1422, 'Engagements': 0, 'Total Installs': 2821, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7596266053953391895/video/7596266053953391895', 'Account': '', 'Received ($)': 2500, 'Estimated ($)': 2500, 'US & EU3 Installs': 495, 'ROW Installs': 847, 'Engagements': 0, 'Total Installs': 1342, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7588452470355987734/video/7588452470355987734', 'Account': '', 'Received ($)': 2500, 'Estimated ($)': 2500, 'US & EU3 Installs': 803, 'ROW Installs': 317, 'Engagements': 0, 'Total Installs': 1120, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7597126976427609366/video/7597126976427609366', 'Account': 'Account 1 (smaller)', 'Received ($)': 2500, 'Estimated ($)': 2500, 'US & EU3 Installs': 2184, 'ROW Installs': 3782, 'Engagements': 0, 'Total Installs': 5966, 'Date First Seen': '2026-02-09'},
    {'TikTok URL': 'https://www.tiktok.com/@7597399641776508163/video/7597399641776508163', 'Account': 'Account 1 (smaller)', 'Received ($)': 2500, 'Estimated ($)': 2500, 'US & EU3 Installs': 717, 'ROW Installs': 1984, 'Engagements': 0, 'Total Installs': 2701, 'Date First Seen': '2026-02-09'},
    {'TikTok URL': 'https://www.tiktok.com/@7597123284848610582/video/7597123284848610582', 'Account': 'Account 1 (smaller)', 'Received ($)': 2500, 'Estimated ($)': 2500, 'US & EU3 Installs': 1196, 'ROW Installs': 1554, 'Engagements': 0, 'Total Installs': 2750, 'Date First Seen': '2026-02-09'},
    {'TikTok URL': 'https://www.tiktok.com/@7597924302243007766/video/7597924302243007766', 'Account': 'Account 1 (smaller)', 'Received ($)': 2500, 'Estimated ($)': 2500, 'US & EU3 Installs': 671, 'ROW Installs': 1269, 'Engagements': 0, 'Total Installs': 1940, 'Date First Seen': '2026-02-09'},
    {'TikTok URL': 'https://www.tiktok.com/@7596261580220763394/video/7596261580220763394', 'Account': '', 'Received ($)': 2233, 'Estimated ($)': 2233, 'US & EU3 Installs': 333, 'ROW Installs': 568, 'Engagements': 0, 'Total Installs': 901, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7600100215269412118/video/7600100215269412118', 'Account': '', 'Received ($)': 2097, 'Estimated ($)': 2097, 'US & EU3 Installs': 303, 'ROW Installs': 582, 'Engagements': 0, 'Total Installs': 885, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7596653525971520790/video/7596653525971520790', 'Account': '', 'Received ($)': 1694, 'Estimated ($)': 1694, 'US & EU3 Installs': 223, 'ROW Installs': 579, 'Engagements': 0, 'Total Installs': 802, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7589641969870081302/video/7589641969870081302', 'Account': '', 'Received ($)': 1638, 'Estimated ($)': 1638, 'US & EU3 Installs': 252, 'ROW Installs': 378, 'Engagements': 0, 'Total Installs': 630, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7596264812846353686/video/7596264812846353686', 'Account': '', 'Received ($)': 1462, 'Estimated ($)': 1462, 'US & EU3 Installs': 209, 'ROW Installs': 417, 'Engagements': 0, 'Total Installs': 626, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7600762883944893718/video/7600762883944893718', 'Account': '', 'Received ($)': 1457, 'Estimated ($)': 1457, 'US & EU3 Installs': 122, 'ROW Installs': 847, 'Engagements': 0, 'Total Installs': 969, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7589953615041711382/video/7589953615041711382', 'Account': '', 'Received ($)': 1386, 'Estimated ($)': 1386, 'US & EU3 Installs': 173, 'ROW Installs': 1, 'Engagements': 0, 'Total Installs': 174, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7597416345443675174/video/7597416345443675174', 'Account': '', 'Received ($)': 1261, 'Estimated ($)': 1261, 'US & EU3 Installs': 154, 'ROW Installs': 491, 'Engagements': 0, 'Total Installs': 645, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7598499844461743382/video/7598499844461743382', 'Account': '', 'Received ($)': 1040, 'Estimated ($)': 1040, 'US & EU3 Installs': 136, 'ROW Installs': 360, 'Engagements': 0, 'Total Installs': 496, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7598144760615947542/video/7598144760615947542', 'Account': '', 'Received ($)': 1018, 'Estimated ($)': 1018, 'US & EU3 Installs': 150, 'ROW Installs': 268, 'Engagements': 0, 'Total Installs': 418, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7598494805294959894/video/7598494805294959894', 'Account': '', 'Received ($)': 893, 'Estimated ($)': 893, 'US & EU3 Installs': 145, 'ROW Installs': 168, 'Engagements': 0, 'Total Installs': 313, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7598604870899289366/video/7598604870899289366', 'Account': '', 'Received ($)': 869, 'Estimated ($)': 869, 'US & EU3 Installs': 152, 'ROW Installs': 109, 'Engagements': 0, 'Total Installs': 261, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7590045885837413654/video/7590045885837413654', 'Account': '', 'Received ($)': 854, 'Estimated ($)': 854, 'US & EU3 Installs': 187, 'ROW Installs': 195, 'Engagements': 0, 'Total Installs': 382, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7596803270442601750/video/7596803270442601750', 'Account': '', 'Received ($)': 818, 'Estimated ($)': 818, 'US & EU3 Installs': 84, 'ROW Installs': 398, 'Engagements': 0, 'Total Installs': 482, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7596800019433295126/video/7596800019433295126', 'Account': '', 'Received ($)': 359, 'Estimated ($)': 359, 'US & EU3 Installs': 63, 'ROW Installs': 44, 'Engagements': 0, 'Total Installs': 107, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7600099102747135254/video/7600099102747135254', 'Account': '', 'Received ($)': 345, 'Estimated ($)': 345, 'US & EU3 Installs': 31, 'ROW Installs': 190, 'Engagements': 0, 'Total Installs': 221, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7597114246433869078/video/7597114246433869078', 'Account': 'Account 1 (smaller)', 'Received ($)': 324, 'Estimated ($)': 324, 'US & EU3 Installs': 49, 'ROW Installs': 79, 'Engagements': 0, 'Total Installs': 128, 'Date First Seen': '2026-02-09'},
    {'TikTok URL': 'https://www.tiktok.com/@7597084269701270806/video/7597084269701270806', 'Account': 'Account 1 (smaller)', 'Received ($)': 314, 'Estimated ($)': 314, 'US & EU3 Installs': 47, 'ROW Installs': 79, 'Engagements': 0, 'Total Installs': 126, 'Date First Seen': '2026-02-09'},
    {'TikTok URL': 'https://www.tiktok.com/@7597737771008085270/video/7597737771008085270', 'Account': '', 'Received ($)': 305, 'Estimated ($)': 305, 'US & EU3 Installs': 53, 'ROW Installs': 40, 'Engagements': 0, 'Total Installs': 93, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7596670194215439638/video/7596670194215439638', 'Account': '', 'Received ($)': 237, 'Estimated ($)': 237, 'US & EU3 Installs': 44, 'ROW Installs': 17, 'Engagements': 0, 'Total Installs': 61, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7600760603992460566/video/7600760603992460566', 'Account': '', 'Received ($)': 230, 'Estimated ($)': 230, 'US & EU3 Installs': 16, 'ROW Installs': 150, 'Engagements': 0, 'Total Installs': 166, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7588890166589246742/video/7588890166589246742', 'Account': '', 'Received ($)': 191, 'Estimated ($)': 191, 'US & EU3 Installs': 38, 'ROW Installs': 1, 'Engagements': 0, 'Total Installs': 39, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7595978340461333782/video/7595978340461333782', 'Account': '', 'Received ($)': 180, 'Estimated ($)': 180, 'US & EU3 Installs': 26, 'ROW Installs': 50, 'Engagements': 0, 'Total Installs': 76, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7589801229585452290/video/7589801229585452290', 'Account': '', 'Received ($)': 120, 'Estimated ($)': 120, 'US & EU3 Installs': 15, 'ROW Installs': 12, 'Engagements': 0, 'Total Installs': 27, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7597805299315068182/video/7597805299315068182', 'Account': 'Account 1 (smaller)', 'Received ($)': 106, 'Estimated ($)': 106, 'US & EU3 Installs': 13, 'ROW Installs': 41, 'Engagements': 0, 'Total Installs': 54, 'Date First Seen': '2026-02-09'},
    {'TikTok URL': 'https://www.tiktok.com/@7585413802045295894/video/7585413802045295894', 'Account': '', 'Received ($)': 84, 'Estimated ($)': 84, 'US & EU3 Installs': 14, 'ROW Installs': 14, 'Engagements': 0, 'Total Installs': 28, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7598924005176921366/video/7598924005176921366', 'Account': '', 'Received ($)': 77, 'Estimated ($)': 77, 'US & EU3 Installs': 10, 'ROW Installs': 27, 'Engagements': 0, 'Total Installs': 37, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7599638833512074518/video/7599638833512074518', 'Account': '', 'Received ($)': 71, 'Estimated ($)': 71, 'US & EU3 Installs': 13, 'ROW Installs': 6, 'Engagements': 0, 'Total Installs': 19, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7534055571922750742/video/7534055571922750742', 'Account': '', 'Received ($)': 66, 'Estimated ($)': 66, 'US & EU3 Installs': 11, 'ROW Installs': 11, 'Engagements': 0, 'Total Installs': 22, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7600491652389752067/video/7600491652389752067', 'Account': '', 'Received ($)': 59, 'Estimated ($)': 59, 'US & EU3 Installs': 11, 'ROW Installs': 4, 'Engagements': 0, 'Total Installs': 15, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7588665170323410198/video/7588665170323410198', 'Account': '', 'Received ($)': 56, 'Estimated ($)': 56, 'US & EU3 Installs': 9, 'ROW Installs': 11, 'Engagements': 0, 'Total Installs': 20, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7600349631934909718/video/7600349631934909718', 'Account': '', 'Received ($)': 53, 'Estimated ($)': 53, 'US & EU3 Installs': 8, 'ROW Installs': 13, 'Engagements': 0, 'Total Installs': 21, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7584815356997373206/video/7584815356997373206', 'Account': '', 'Received ($)': 35, 'Estimated ($)': 35, 'US & EU3 Installs': 7, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 7, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7599639960802659606/video/7599639960802659606', 'Account': '', 'Received ($)': 25, 'Estimated ($)': 25, 'US & EU3 Installs': 4, 'ROW Installs': 5, 'Engagements': 0, 'Total Installs': 9, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7600069960731200790/video/7600069960731200790', 'Account': '', 'Received ($)': 19, 'Estimated ($)': 19, 'US & EU3 Installs': 2, 'ROW Installs': 9, 'Engagements': 0, 'Total Installs': 11, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7588593850516589846/video/7588593850516589846', 'Account': '', 'Received ($)': 18, 'Estimated ($)': 18, 'US & EU3 Installs': 3, 'ROW Installs': 3, 'Engagements': 0, 'Total Installs': 6, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7599750339352595735/video/7599750339352595735', 'Account': '', 'Received ($)': 16, 'Estimated ($)': 16, 'US & EU3 Installs': 2, 'ROW Installs': 6, 'Engagements': 0, 'Total Installs': 8, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7584827008618646806/video/7584827008618646806', 'Account': '', 'Received ($)': 15, 'Estimated ($)': 15, 'US & EU3 Installs': 2, 'ROW Installs': 5, 'Engagements': 0, 'Total Installs': 7, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7592963003901840662/video/7592963003901840662', 'Account': '', 'Received ($)': 15, 'Estimated ($)': 15, 'US & EU3 Installs': 2, 'ROW Installs': 5, 'Engagements': 0, 'Total Installs': 7, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7538858594683440406/video/7538858594683440406', 'Account': '', 'Received ($)': 14, 'Estimated ($)': 14, 'US & EU3 Installs': 2, 'ROW Installs': 4, 'Engagements': 0, 'Total Installs': 6, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7584811033206803734/video/7584811033206803734', 'Account': '', 'Received ($)': 13, 'Estimated ($)': 13, 'US & EU3 Installs': 2, 'ROW Installs': 3, 'Engagements': 0, 'Total Installs': 5, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7595235096404266262/video/7595235096404266262', 'Account': '', 'Received ($)': 11, 'Estimated ($)': 11, 'US & EU3 Installs': 2, 'ROW Installs': 1, 'Engagements': 0, 'Total Installs': 3, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7589183622859558167/video/7589183622859558167', 'Account': '', 'Received ($)': 11, 'Estimated ($)': 11, 'US & EU3 Installs': 2, 'ROW Installs': 1, 'Engagements': 0, 'Total Installs': 3, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7600770165768228118/video/7600770165768228118', 'Account': '', 'Received ($)': 8, 'Estimated ($)': 8, 'US & EU3 Installs': 1, 'ROW Installs': 3, 'Engagements': 0, 'Total Installs': 4, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7599275177226046742/video/7599275177226046742', 'Account': '', 'Received ($)': 8, 'Estimated ($)': 8, 'US & EU3 Installs': 1, 'ROW Installs': 3, 'Engagements': 0, 'Total Installs': 4, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7596766270058417430/video/7596766270058417430', 'Account': '', 'Received ($)': 7, 'Estimated ($)': 7, 'US & EU3 Installs': 1, 'ROW Installs': 2, 'Engagements': 0, 'Total Installs': 3, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7599177144627154179/video/7599177144627154179', 'Account': '', 'Received ($)': 6, 'Estimated ($)': 6, 'US & EU3 Installs': 1, 'ROW Installs': 1, 'Engagements': 0, 'Total Installs': 2, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7589643666491297046/video/7589643666491297046', 'Account': '', 'Received ($)': 6, 'Estimated ($)': 6, 'US & EU3 Installs': 1, 'ROW Installs': 1, 'Engagements': 0, 'Total Installs': 2, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7590120022803205397/video/7590120022803205397', 'Account': '', 'Received ($)': 5, 'Estimated ($)': 5, 'US & EU3 Installs': 0, 'ROW Installs': 5, 'Engagements': 0, 'Total Installs': 5, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7592645042695802135/video/7592645042695802135', 'Account': '', 'Received ($)': 5, 'Estimated ($)': 5, 'US & EU3 Installs': 1, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 1, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7599026448749382934/video/7599026448749382934', 'Account': '', 'Received ($)': 5, 'Estimated ($)': 5, 'US & EU3 Installs': 1, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 1, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7586604785374104854/video/7586604785374104854', 'Account': '', 'Received ($)': 5, 'Estimated ($)': 5, 'US & EU3 Installs': 1, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 1, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7597362819398503702/video/7597362819398503702', 'Account': 'Account 1 (smaller)', 'Received ($)': 5, 'Estimated ($)': 5, 'US & EU3 Installs': 1, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 1, 'Date First Seen': '2026-02-09'},
    {'TikTok URL': 'https://www.tiktok.com/@7599372602066390294/video/7599372602066390294', 'Account': '', 'Received ($)': 3, 'Estimated ($)': 3, 'US & EU3 Installs': 0, 'ROW Installs': 3, 'Engagements': 0, 'Total Installs': 3, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7599371798970371350/video/7599371798970371350', 'Account': '', 'Received ($)': 3, 'Estimated ($)': 3, 'US & EU3 Installs': 0, 'ROW Installs': 3, 'Engagements': 0, 'Total Installs': 3, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7599032518100241686/video/7599032518100241686', 'Account': 'Account 1 (smaller)', 'Received ($)': 3, 'Estimated ($)': 3, 'US & EU3 Installs': 0, 'ROW Installs': 3, 'Engagements': 0, 'Total Installs': 3, 'Date First Seen': '2026-02-09'},
    {'TikTok URL': 'https://www.tiktok.com/@7598241192240778518/video/7598241192240778518', 'Account': '', 'Received ($)': 2, 'Estimated ($)': 2, 'US & EU3 Installs': 0, 'ROW Installs': 2, 'Engagements': 0, 'Total Installs': 2, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7589955663128775958/video/7589955663128775958', 'Account': '', 'Received ($)': 2, 'Estimated ($)': 2, 'US & EU3 Installs': 1, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 1, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7592952482741964054/video/7592952482741964054', 'Account': '', 'Received ($)': 2, 'Estimated ($)': 2, 'US & EU3 Installs': 0, 'ROW Installs': 2, 'Engagements': 0, 'Total Installs': 2, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7590115654942854422/video/7590115654942854422', 'Account': '', 'Received ($)': 2, 'Estimated ($)': 2, 'US & EU3 Installs': 0, 'ROW Installs': 2, 'Engagements': 0, 'Total Installs': 2, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7598188911415938326/video/7598188911415938326', 'Account': '', 'Received ($)': 2, 'Estimated ($)': 2, 'US & EU3 Installs': 0, 'ROW Installs': 2, 'Engagements': 0, 'Total Installs': 2, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7585524986576637186/video/7585524986576637186', 'Account': '', 'Received ($)': 2, 'Estimated ($)': 2, 'US & EU3 Installs': 0, 'ROW Installs': 2, 'Engagements': 0, 'Total Installs': 2, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7588896788531252502/video/7588896788531252502', 'Account': '', 'Received ($)': 2, 'Estimated ($)': 2, 'US & EU3 Installs': 0, 'ROW Installs': 2, 'Engagements': 0, 'Total Installs': 2, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7597064708012920067/video/7597064708012920067', 'Account': 'Account 1 (smaller)', 'Received ($)': 2, 'Estimated ($)': 2, 'US & EU3 Installs': 0, 'ROW Installs': 2, 'Engagements': 0, 'Total Installs': 2, 'Date First Seen': '2026-02-09'},
    {'TikTok URL': 'https://www.tiktok.com/@7599268989629549827/video/7599268989629549827', 'Account': '', 'Received ($)': 1, 'Estimated ($)': 1, 'US & EU3 Installs': 0, 'ROW Installs': 1, 'Engagements': 0, 'Total Installs': 1, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7589294163988974870/video/7589294163988974870', 'Account': '', 'Received ($)': 1, 'Estimated ($)': 1, 'US & EU3 Installs': 0, 'ROW Installs': 1, 'Engagements': 0, 'Total Installs': 1, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7595964773100014870/video/7595964773100014870', 'Account': '', 'Received ($)': 1, 'Estimated ($)': 1, 'US & EU3 Installs': 0, 'ROW Installs': 1, 'Engagements': 0, 'Total Installs': 1, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7600781841251061014/video/7600781841251061014', 'Account': '', 'Received ($)': 1, 'Estimated ($)': 1, 'US & EU3 Installs': 0, 'ROW Installs': 1, 'Engagements': 0, 'Total Installs': 1, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7600362027797826838/video/7600362027797826838', 'Account': '', 'Received ($)': 1, 'Estimated ($)': 1, 'US & EU3 Installs': 0, 'ROW Installs': 1, 'Engagements': 0, 'Total Installs': 1, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7592911940830219542/video/7592911940830219542', 'Account': '', 'Received ($)': 1, 'Estimated ($)': 1, 'US & EU3 Installs': 0, 'ROW Installs': 1, 'Engagements': 0, 'Total Installs': 1, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7595908644831694082/video/7595908644831694082', 'Account': '', 'Received ($)': 1, 'Estimated ($)': 1, 'US & EU3 Installs': 0, 'ROW Installs': 1, 'Engagements': 0, 'Total Installs': 1, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7601487341189975318/video/7601487341189975318', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 1539, 'US & EU3 Installs': 284, 'ROW Installs': 119, 'Engagements': 0, 'Total Installs': 403, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7601835249730915606/video/7601835249730915606', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 56, 'US & EU3 Installs': 9, 'ROW Installs': 11, 'Engagements': 0, 'Total Installs': 20, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7601565353419853078/video/7601565353419853078', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 196, 'US & EU3 Installs': 37, 'ROW Installs': 11, 'Engagements': 0, 'Total Installs': 48, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7604896624334228758/video/7604896624334228758', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 3, 'US & EU3 Installs': 0, 'ROW Installs': 3, 'Engagements': 0, 'Total Installs': 3, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7601828002359430422/video/7601828002359430422', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 78, 'US & EU3 Installs': 15, 'ROW Installs': 3, 'Engagements': 0, 'Total Installs': 18, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7603420165602037014/video/7603420165602037014', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 128, 'US & EU3 Installs': 23, 'ROW Installs': 13, 'Engagements': 0, 'Total Installs': 36, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7530652618117582102/video/7530652618117582102', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7601622544600354070/video/7601622544600354070', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 5, 'US & EU3 Installs': 1, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 1, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7589958196710722838/video/7589958196710722838', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7601242576053652758/video/7601242576053652758', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7601537953378143510/video/7601537953378143510', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 179, 'US & EU3 Installs': 33, 'ROW Installs': 14, 'Engagements': 0, 'Total Installs': 47, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7601470612418170135/video/7601470612418170135', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 6, 'US & EU3 Installs': 1, 'ROW Installs': 1, 'Engagements': 0, 'Total Installs': 2, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7601632405601996054/video/7601632405601996054', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 6, 'US & EU3 Installs': 1, 'ROW Installs': 1, 'Engagements': 0, 'Total Installs': 2, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7596811565123374358/video/7596811565123374358', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7604153402934725890/video/7604153402934725890', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 3, 'US & EU3 Installs': 0, 'ROW Installs': 3, 'Engagements': 0, 'Total Installs': 3, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7594456162293026070/video/7594456162293026070', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7603017097614232854/video/7603017097614232854', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 5, 'US & EU3 Installs': 1, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 1, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7595560825469930774/video/7595560825469930774', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7590078610845879574/video/7590078610845879574', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7601468586128952598/video/7601468586128952598', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7602393658146524418/video/7602393658146524418', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 10, 'US & EU3 Installs': 2, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 2, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7598278879643569430/video/7598278879643569430', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7598212457932393750/video/7598212457932393750', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7599172358380719382/video/7599172358380719382', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7601614140414708995/video/7601614140414708995', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 18, 'US & EU3 Installs': 3, 'ROW Installs': 3, 'Engagements': 0, 'Total Installs': 6, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7590381156978199831/video/7590381156978199831', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7598963080218873111/video/7598963080218873111', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7603374540235803926/video/7603374540235803926', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 81, 'US & EU3 Installs': 6, 'ROW Installs': 51, 'Engagements': 0, 'Total Installs': 57, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7530023517908159766/video/7530023517908159766', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7601515548182138134/video/7601515548182138134', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 28, 'US & EU3 Installs': 5, 'ROW Installs': 3, 'Engagements': 0, 'Total Installs': 8, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7601611906779974934/video/7601611906779974934', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 21, 'US & EU3 Installs': 4, 'ROW Installs': 1, 'Engagements': 0, 'Total Installs': 5, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7528031035309821206/video/7528031035309821206', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7603359664142568726/video/7603359664142568726', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7601167547966164246/video/7601167547966164246', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 28, 'US & EU3 Installs': 2, 'ROW Installs': 18, 'Engagements': 0, 'Total Installs': 20, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7584426482357882115/video/7584426482357882115', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7601477473322159382/video/7601477473322159382', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 5, 'US & EU3 Installs': 1, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 1, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7599666081636322583/video/7599666081636322583', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7598945928455720214/video/7598945928455720214', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7603101081257790742/video/7603101081257790742', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 8, 'US & EU3 Installs': 1, 'ROW Installs': 3, 'Engagements': 0, 'Total Installs': 4, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7599276140229856534/video/7599276140229856534', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7601245488603565334/video/7601245488603565334', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 20, 'US & EU3 Installs': 4, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 4, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7601836046355074326/video/7601836046355074326', 'Account': '', 'Received ($)': 0, 'Estimated ($)': 7, 'US & EU3 Installs': 0, 'ROW Installs': 7, 'Engagements': 0, 'Total Installs': 7, 'Date First Seen': '2026-02-14'},
    {'TikTok URL': 'https://www.tiktok.com/@7597795533490507030/video/7597795533490507030', 'Account': 'Account 1 (smaller)', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-09'},
    {'TikTok URL': 'https://www.tiktok.com/@7597363566458539286/video/7597363566458539286', 'Account': 'Account 1 (smaller)', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-09'},
    {'TikTok URL': 'https://www.tiktok.com/@7597171707987709206/video/7597171707987709206', 'Account': 'Account 1 (smaller)', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-09'},
    {'TikTok URL': 'https://www.tiktok.com/@7591292533720878358/video/7591292533720878358', 'Account': 'Account 1 (smaller)', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-09'},
    {'TikTok URL': 'https://www.tiktok.com/@7597796013792939286/video/7597796013792939286', 'Account': 'Account 1 (smaller)', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-09'},
    {'TikTok URL': 'https://www.tiktok.com/@7597800531087658262/video/7597800531087658262', 'Account': 'Account 1 (smaller)', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-09'},
    {'TikTok URL': 'https://www.tiktok.com/@7597080404323028246/video/7597080404323028246', 'Account': 'Account 1 (smaller)', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-09'},
    {'TikTok URL': 'https://www.tiktok.com/@7597108883814944022/video/7597108883814944022', 'Account': 'Account 1 (smaller)', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-09'},
    {'TikTok URL': 'https://www.tiktok.com/@7603100212176997654/video/7603100212176997654', 'Account': 'Account 1 (smaller)', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-09'},
    {'TikTok URL': 'https://www.tiktok.com/@7597012528090107158/video/7597012528090107158', 'Account': 'Account 1 (smaller)', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-09'},
    {'TikTok URL': 'https://www.tiktok.com/@7597801729450577174/video/7597801729450577174', 'Account': 'Account 1 (smaller)', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-09'},
    {'TikTok URL': 'https://www.tiktok.com/@7597364619035888899/video/7597364619035888899', 'Account': 'Account 1 (smaller)', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-09'},
    {'TikTok URL': 'https://www.tiktok.com/@7590877569394773270/video/7590877569394773270', 'Account': 'Account 1 (smaller)', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-09'},
    {'TikTok URL': 'https://www.tiktok.com/@7582151219729272086/video/7582151219729272086', 'Account': 'Account 1 (smaller)', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-09'},
    {'TikTok URL': 'https://www.tiktok.com/@7581890594608336150/video/7581890594608336150', 'Account': 'Account 1 (smaller)', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-09'},
    {'TikTok URL': 'https://www.tiktok.com/@7597928583667010819/video/7597928583667010819', 'Account': 'Account 1 (smaller)', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-09'},
    {'TikTok URL': 'https://www.tiktok.com/@7582190483133238550/video/7582190483133238550', 'Account': 'Account 1 (smaller)', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-09'},
    {'TikTok URL': 'https://www.tiktok.com/@7581937222518115606/video/7581937222518115606', 'Account': 'Account 1 (smaller)', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-09'},
    {'TikTok URL': 'https://www.tiktok.com/@capcuttemplatesai/video/7433474060589223201', 'Account': 'capcuttemplatesai', 'Received ($)': 0, 'Estimated ($)': 0, 'US & EU3 Installs': 0, 'ROW Installs': 0, 'Engagements': 0, 'Total Installs': 0, 'Date First Seen': '2026-02-09'},
]


# =============================================================================
# VARIANT ALLOCATION & STOP RULES (v3.5.0 Option B - Baked In)
# =============================================================================

VARIANT_CACHE_DEFAULT_TTL = 7  # days

def _strip_emoji(s: str) -> str:
    """Strip leading emoji and whitespace from action_window/trajectory values."""
    if not s:
        return ""
    return _re.sub(r'^[^\x00-\x7F]+\s*', '', str(s)).strip()


def _as_float_vel(x) -> Optional[float]:
    """Robust float parsing for '+8,747/day', '32.9h', 8747, etc."""
    if x is None:
        return None
    try:
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x).strip().lower()
        if not s:
            return None
        s = s.replace(",", "").replace("/day", "").replace("per day", "").replace("h", "")
        if s.startswith("+"):
            s = s[1:]
        return float(s)
    except Exception:
        return None


def calc_recommended_variants(aw: str, tr: str, age: Optional[float], cur: Optional[float]) -> int:
    """Calculate how many template variants to build (0/1/2/3/5/7)."""
    awu = _strip_emoji(aw).upper()
    tru = _strip_emoji(tr).upper()
    agev = age if age is not None else 999999.0
    curv = cur if cur is not None else 0.0

    # Hard stop zones
    if awu in ("PEAKED", "TOO LATE", "WINDOW CLOSING"):
        return 0
    if tru in ("DECLINING", "CRASHING"):
        return 0
    if agev >= 72:
        return 0

    # Last-chance tier (60-72h): exceptional only
    if 60 <= agev < 72:
        if awu in ("ACT NOW", "6-12H") and tru in ("EXPLOSIVE", "STRONG") and curv >= 5000:
            return 1
        return 0

    # Normal allocation
    if awu == "ACT NOW" and tru == "EXPLOSIVE" and agev <= 24:
        return 7
    if awu == "ACT NOW" and tru in ("EXPLOSIVE", "STRONG"):
        return 5
    if awu == "ACT NOW" and tru == "MODERATE":
        return 3
    if awu == "6-12H" and tru == "EXPLOSIVE":
        return 5
    if awu == "6-12H" and tru in ("STRONG", "MODERATE"):
        return 3
    if awu == "12-24H" and tru == "STRONG":
        return 3
    if awu == "12-24H" and tru == "MODERATE":
        return 2
    if awu == "12-24H" and tru == "FLAT":
        return 1
    return 0


def calc_stop_building(aw: str, tr: str, age: Optional[float], streak: int) -> Tuple[bool, str]:
    """Determine if building should stop and why."""
    awu = _strip_emoji(aw).upper()
    tru = _strip_emoji(tr).upper()
    agev = age if age is not None else 999999.0

    if tru in ("DECLINING", "CRASHING"):
        return True, "DECLINING_TRAJECTORY"
    if awu in ("PEAKED", "TOO LATE"):
        return True, "WINDOW_OVER"
    if agev >= 72:
        return True, "AGE_OVER_72H"
    if streak >= 2:
        return True, "VELOCITY_NONPOS_2_RUNS"
    return False, ""


def load_streak_cache(path: str) -> Dict:
    """Load velocity streak cache from JSON file."""
    if not path or not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            raw = json.load(f)
        out = {}
        for url, obj in raw.items():
            if isinstance(obj, dict):
                out[str(url)] = {"streak": int(obj.get("streak", 0) or 0), "last_seen": obj.get("last_seen")}
            else:
                out[str(url)] = {"streak": int(obj) if str(obj).isdigit() else 0, "last_seen": None}
        return out
    except Exception:
        return {}


def prune_streak_cache(cache: dict, ttl_days: int) -> dict:
    """Remove entries older than TTL days."""
    if ttl_days <= 0:
        return cache
    from datetime import date as _date
    cutoff = _date.today() - timedelta(days=ttl_days)
    keep = {}
    for url, obj in cache.items():
        ls = obj.get("last_seen")
        if not ls:
            keep[url] = obj
            continue
        try:
            d = datetime.strptime(ls, "%Y-%m-%d").date()
            if d >= cutoff:
                keep[url] = obj
        except Exception:
            keep[url] = obj
    return keep


def save_streak_cache(path: str, cache: dict) -> None:
    """Save velocity streak cache to JSON file."""
    if not path:
        return
    with open(path, "w", encoding="utf-8") as f:
        json.dump(cache, f, indent=2, sort_keys=True)


# =============================================================================
# VELOCITY PREDICTION ENGINE
# =============================================================================

@dataclass
class VelocityPrediction:
    """Holds prediction data for a trend"""
    current_momentum: float
    velocity: float  # momentum change per day
    acceleration: float  # velocity change per day
    predicted_6h: float
    predicted_12h: float
    predicted_24h: float
    peak_estimate_hours: Optional[float]  # estimated hours until peak (None if declining)
    trajectory: str  # EXPLOSIVE, STRONG, MODERATE, WEAK, DECLINING, CRASHING
    confidence: str  # HIGH, MEDIUM, LOW based on data quality
    action_window: str  # "ACT NOW", "6-12H", "12-24H", "MONITOR", "TOO LATE"


def calculate_velocity_predictions(
    df_today: pd.DataFrame,
    df_yesterday: pd.DataFrame = None,
    df_2days_ago: pd.DataFrame = None
) -> pd.DataFrame:
    """
    Calculate velocity-based predictions for all trends.
    
    Uses up to 3 days of data to calculate velocity and acceleration.
    Falls back gracefully when historical data is missing.
    """
    df = df_today.copy()
    
    # Auto-calculate metrics if missing (raw Apify data)
    df = _ensure_calculated_metrics(df)
    if df_yesterday is not None:
        df_yesterday = _ensure_calculated_metrics(df_yesterday)
    if df_2days_ago is not None:
        df_2days_ago = _ensure_calculated_metrics(df_2days_ago)
    
    # Merge with yesterday's data if available
    if df_yesterday is not None and len(df_yesterday) > 0:
        yest_dedup = df_yesterday.drop_duplicates(subset=['webVideoUrl'], keep='first')
        yesterday_momentum = yest_dedup[['webVideoUrl', 'momentum_score']].copy()
        yesterday_momentum.columns = ['webVideoUrl', 'momentum_yesterday']
        df = df.merge(yesterday_momentum, on='webVideoUrl', how='left')
    else:
        df['momentum_yesterday'] = np.nan
    
    # Merge with 2-days-ago data if available
    if df_2days_ago is not None and len(df_2days_ago) > 0:
        d2_dedup = df_2days_ago.drop_duplicates(subset=['webVideoUrl'], keep='first')
        old_momentum = d2_dedup[['webVideoUrl', 'momentum_score']].copy()
        old_momentum.columns = ['webVideoUrl', 'momentum_2days']
        df = df.merge(old_momentum, on='webVideoUrl', how='left')
    else:
        df['momentum_2days'] = np.nan
    
    # Calculate velocity (change per day)
    df['velocity'] = df['momentum_score'] - df['momentum_yesterday'].fillna(df['momentum_score'])
    
    # Calculate acceleration (change in velocity)
    if df_yesterday is not None and df_2days_ago is not None:
        velocity_yesterday = df['momentum_yesterday'] - df['momentum_2days']
        df['acceleration'] = df['velocity'] - velocity_yesterday.fillna(0)
    else:
        df['acceleration'] = 0
    
    # Predict future momentum using physics model: position + velocity*t + 0.5*acceleration*t^2
    # But cap acceleration impact to avoid runaway predictions
    df['acceleration_capped'] = df['acceleration'].clip(-50, 50)
    
    # Predictions (time in days: 0.25 = 6h, 0.5 = 12h, 1.0 = 24h)
    df['predicted_6h'] = (
        df['momentum_score'] + 
        df['velocity'] * 0.25 + 
        0.5 * df['acceleration_capped'] * (0.25 ** 2)
    ).clip(lower=0)
    
    df['predicted_12h'] = (
        df['momentum_score'] + 
        df['velocity'] * 0.5 + 
        0.5 * df['acceleration_capped'] * (0.5 ** 2)
    ).clip(lower=0)
    
    df['predicted_24h'] = (
        df['momentum_score'] + 
        df['velocity'] * 1.0 + 
        0.5 * df['acceleration_capped'] * (1.0 ** 2)
    ).clip(lower=0)
    
    # Determine trajectory
    def get_trajectory(velocity):
        if velocity >= VELOCITY_THRESHOLDS['EXPLOSIVE']:
            return '🚀 EXPLOSIVE'
        elif velocity >= VELOCITY_THRESHOLDS['STRONG']:
            return '📈 STRONG'
        elif velocity >= VELOCITY_THRESHOLDS['MODERATE']:
            return '↗️ MODERATE'
        elif velocity >= VELOCITY_THRESHOLDS['WEAK']:
            return '➡️ FLAT'
        elif velocity >= VELOCITY_THRESHOLDS['DECLINING']:
            return '↘️ DECLINING'
        else:
            return '📉 CRASHING'
    
    df['trajectory'] = df['velocity'].apply(get_trajectory)
    
    # Estimate peak timing (when velocity will hit 0)
    # peak_time = -velocity / acceleration (only valid if acceleration < 0 and velocity > 0)
    def estimate_peak(row):
        if row['acceleration'] < -5 and row['velocity'] > 0:
            peak_hours = (-row['velocity'] / row['acceleration']) * 24
            if 0 < peak_hours < 72:
                return round(peak_hours, 1)
        return None
    
    df['peak_estimate_hours'] = df.apply(estimate_peak, axis=1)
    
    # Determine confidence based on data availability
    def get_confidence(row):
        if pd.notna(row.get('momentum_2days')) and pd.notna(row.get('momentum_yesterday')):
            return 'HIGH'
        elif pd.notna(row.get('momentum_yesterday')):
            return 'MEDIUM'
        else:
            return 'LOW'
    
    df['prediction_confidence'] = df.apply(get_confidence, axis=1)
    
    # Determine action window
    def get_action_window(row):
        age = row.get('age_hours', 999)
        velocity = row['velocity']
        momentum = row['momentum_score']
        predicted_24h = row['predicted_24h']
        has_velocity_data = pd.notna(row.get('momentum_yesterday')) and row.get('momentum_yesterday') != row.get('momentum_score', 0)
        
        # Too old - 72h window closing
        if age > 60:
            return '⚠️ WINDOW CLOSING'
        
        # If we have real velocity data, use it
        if has_velocity_data:
            # Explosive growth - act immediately
            if velocity >= 200 and momentum >= 1000:
                return '🔴 ACT NOW'
            
            # Strong growth - act within 6-12h
            if velocity >= 100 and momentum >= 500:
                return '🟠 6-12H'
            
            # Moderate growth but predicted to be big
            if velocity >= 50 and predicted_24h >= 2000:
                return '🟡 12-24H'
            
            # Flat or declining
            if velocity <= 0:
                if momentum >= 2000:
                    return '⚠️ PEAKED'
                else:
                    return '❌ TOO LATE'
            
            # Default - worth monitoring
            return '🟢 MONITOR'
        
        else:
            # No velocity data — classify on momentum + age alone
            if momentum >= 3000 and age <= 24:
                return '🔴 ACT NOW'
            if momentum >= 2000 and age <= 36:
                return '🟠 6-12H'
            if momentum >= 1000 and age <= 48:
                return '🟡 12-24H'
            if momentum >= 500:
                return '🟢 MONITOR'
            return '❌ TOO LATE'
    
    df['action_window'] = df.apply(get_action_window, axis=1)
    
    return df


def _ensure_calculated_metrics(df: pd.DataFrame) -> pd.DataFrame:
    """
    Ensure momentum_score, age_hours, shares_per_hour etc. exist.
    Calculate from raw Apify fields if missing.
    """
    df = df.copy()
    
    # Calculate age_hours if missing
    if 'age_hours' not in df.columns or df['age_hours'].isna().all():
        if 'createTimeISO' in df.columns:
            now = pd.Timestamp.now(tz='UTC')
            df['createTimeISO'] = pd.to_datetime(df['createTimeISO'], utc=True, errors='coerce')
            df['age_hours'] = (now - df['createTimeISO']).dt.total_seconds() / 3600
            df['age_hours'] = df['age_hours'].clip(lower=0.1)  # Prevent division by zero
        else:
            df['age_hours'] = 24  # Default assumption
    
    # Ensure numeric columns exist
    for col in ['shareCount', 'diggCount', 'playCount']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    # Force numeric type on metric columns (may arrive as strings from some sources)
    for col in ['age_hours', 'momentum_score', 'shares_per_hour', 'likes_per_hour', 'views_per_hour']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    # Calculate per-hour metrics if missing
    if 'shares_per_hour' not in df.columns or df['shares_per_hour'].isna().all():
        if 'shareCount' in df.columns and 'age_hours' in df.columns:
            df['shares_per_hour'] = df['shareCount'] / df['age_hours'].clip(lower=0.1)
    
    if 'likes_per_hour' not in df.columns or df['likes_per_hour'].isna().all():
        if 'diggCount' in df.columns and 'age_hours' in df.columns:
            df['likes_per_hour'] = df['diggCount'] / df['age_hours'].clip(lower=0.1)
    
    if 'views_per_hour' not in df.columns or df['views_per_hour'].isna().all():
        if 'playCount' in df.columns and 'age_hours' in df.columns:
            df['views_per_hour'] = df['playCount'] / df['age_hours'].clip(lower=0.1)
    
    # Calculate momentum_score if missing
    if 'momentum_score' not in df.columns or df['momentum_score'].isna().all():
        shares_h = df.get('shares_per_hour', pd.Series(0, index=df.index))
        likes_h = df.get('likes_per_hour', pd.Series(0, index=df.index))
        views_h = df.get('views_per_hour', pd.Series(0, index=df.index))
        
        df['momentum_score'] = (
            shares_h.fillna(0) * 10 + 
            likes_h.fillna(0) * 3 + 
            views_h.fillna(0) * 0.01
        )
    
    # Ensure Market column exists
    if 'Market' not in df.columns:
        df['Market'] = '🇬🇧 UK ONLY'  # Default
    
    # Ensure AI_CATEGORY exists
    if 'AI_CATEGORY' not in df.columns:
        df['AI_CATEGORY'] = 'Unknown'
    
    return df


def create_velocity_summary(df: pd.DataFrame, cache_path: str = None) -> pd.DataFrame:
    """Create a summary view of velocity predictions sorted by opportunity.
    
    Now includes variant allocation and stop rules (v3.5.0 Option B).
    """
    
    cols = [
        'webVideoUrl', 'text', 'author', 'age_hours',
        'momentum_score', 'velocity', 'acceleration',
        'predicted_6h', 'predicted_12h', 'predicted_24h',
        'trajectory', 'peak_estimate_hours', 'prediction_confidence',
        'action_window', 'Market', 'AI_CATEGORY'
    ]
    
    # Filter to available columns
    available_cols = [c for c in cols if c in df.columns]
    summary = df[available_cols].copy()
    
    # Create opportunity score for sorting (with NaN handling)
    # Prioritize: high predicted growth + young age + ACT NOW status
    summary['opportunity_score'] = (
        (summary['predicted_24h'].fillna(0) - summary['momentum_score'].fillna(0)) * 0.5 +  # Growth potential
        summary['velocity'].fillna(0) * 0.3 +  # Current velocity
        (72 - summary['age_hours'].fillna(72).clip(upper=72)) * 10  # Youth bonus
    )
    
    # Boost ACT NOW items
    summary.loc[summary['action_window'].str.contains('ACT NOW', na=False), 'opportunity_score'] *= 1.5
    
    # Sort by opportunity score
    summary = summary.sort_values('opportunity_score', ascending=False)
    
    # Format for display (with NaN handling)
    summary['Trend'] = summary['text'].fillna('').astype(str).str[:60] + '...' if 'text' in summary.columns else ''
    summary['Creator'] = summary.get('author', '').fillna('')
    summary['Age'] = summary['age_hours'].fillna(0).apply(lambda x: f"{x:.1f}h")
    summary['Current'] = summary['momentum_score'].fillna(0).astype(int)
    summary['Velocity'] = summary['velocity'].fillna(0).apply(lambda x: f"{x:+.0f}/day")
    summary['In 6h'] = summary['predicted_6h'].fillna(0).astype(int)
    summary['In 12h'] = summary['predicted_12h'].fillna(0).astype(int)
    summary['In 24h'] = summary['predicted_24h'].fillna(0).astype(int)
    summary['Peak In'] = summary['peak_estimate_hours'].apply(
        lambda x: f"{x:.0f}h" if pd.notna(x) else "N/A"
    )
    
    # ── Variant allocation & stop rules (Option B inline) ──
    cache = prune_streak_cache(load_streak_cache(cache_path or ''), VARIANT_CACHE_DEFAULT_TTL)
    today_str = datetime.now().strftime("%Y-%m-%d")
    
    rec_variants = []
    streaks = []
    stops = []
    stop_reasons = []
    
    for _, row in summary.iterrows():
        aw = str(row.get('action_window', ''))
        tr = str(row.get('trajectory', ''))
        age_val = row.get('age_hours', None)
        cur_val = row.get('momentum_score', None)
        vel_val = row.get('velocity', None)
        url = str(row.get('webVideoUrl', ''))
        
        # Streak calculation
        prev = cache.get(url, {"streak": 0, "last_seen": None})
        prev_streak = int(prev.get("streak", 0) or 0)
        
        vel_float = _as_float_vel(vel_val)
        if vel_float is not None and vel_float <= 0:
            streak = prev_streak + 1
        elif vel_float is not None and vel_float > 0:
            streak = 0
        else:
            streak = prev_streak  # unknown velocity -> hold
        
        if url:
            cache[url] = {"streak": streak, "last_seen": today_str}
        
        rv = calc_recommended_variants(aw, tr, age_val, cur_val)
        stop, reason = calc_stop_building(aw, tr, age_val, streak)
        
        rec_variants.append(rv)
        streaks.append(streak)
        stops.append(stop)
        stop_reasons.append(reason)
    
    summary['recommended_variants'] = rec_variants
    summary['velocity_nonpos_streak'] = streaks
    summary['stop_building'] = stops
    summary['stop_reason'] = stop_reasons
    
    # Save updated cache
    save_streak_cache(cache_path or '', cache)
    
    # Final columns for output
    output_cols = [
        'action_window', 'trajectory', 'Trend', 'Creator', 'Age',
        'Current', 'Velocity', 'In 6h', 'In 12h', 'In 24h',
        'Peak In', 'prediction_confidence', 'Market', 'webVideoUrl',
        'recommended_variants', 'velocity_nonpos_streak', 'stop_building', 'stop_reason'
    ]
    
    return summary[[c for c in output_cols if c in summary.columns]]


# =============================================================================
# COMPETITOR ANALYSIS ENGINE
# =============================================================================

@dataclass
class CompetitorInsight:
    """Analysis of competitor behavior vs yours"""
    trend_url: str
    trend_text: str
    competitor_posted: bool
    you_posted: bool
    momentum_when_competitor_posted: float
    current_momentum: float
    gap_type: str  # "MISSED_OPPORTUNITY", "BEAT_THEM", "BOTH_CAUGHT", "NEITHER"
    hours_behind: Optional[float]  # How many hours after competitor you posted (None if you didn't)
    potential_revenue_missed: float  # Estimated £ based on momentum


def analyze_competitor_gaps(
    df_today: pd.DataFrame,
    df_historical: pd.DataFrame = None,  # Last 7 days aggregated
    your_accounts: List[str] = YOUR_ACCOUNTS,
    competitor_accounts: List[str] = COMPETITOR_ACCOUNTS
) -> pd.DataFrame:
    """
    Analyze what trends the competitor catches that you miss.
    
    Returns DataFrame with gap analysis.
    """
    df = df_today.copy()
    
    # Identify posts by you vs competitor
    df['is_yours'] = df['author'].isin(your_accounts)
    df['is_competitor'] = df['author'].isin(competitor_accounts)
    
    # Group by trend pattern (using text similarity would be better, but URL works for exact matches)
    # For now, identify trends where competitor posted but you didn't
    
    competitor_urls = set(df[df['is_competitor']]['webVideoUrl'])
    your_urls = set(df[df['is_yours']]['webVideoUrl'])
    
    # Find high-momentum trends competitor caught
    competitor_trends = df[df['is_competitor']].copy()
    
    # Analyze each competitor post
    results = []
    
    for _, comp_row in competitor_trends.iterrows():
        # Did you also post this trend?
        you_also_posted = comp_row['webVideoUrl'] in your_urls
        
        # Get your version if exists
        your_version = df[(df['webVideoUrl'] == comp_row['webVideoUrl']) & df['is_yours']]
        
        if len(your_version) > 0:
            gap_type = 'BOTH_CAUGHT'
            hours_behind = (
                your_version.iloc[0]['age_hours'] - comp_row['age_hours']
            ) if 'age_hours' in your_version.columns else None
        else:
            gap_type = 'MISSED_BY_YOU'
            hours_behind = None
        
        # Estimate missed revenue using data-driven model
        from revenue_model import estimate_competitor_revenue as _est_comp
        _est = _est_comp(comp_row['momentum_score'], comp_row.get('shares_per_hour'), comp_row.get('age_hours'))
        potential_missed = _est['estimated_revenue'] if not you_also_posted else 0
        
        results.append({
            'trend_url': comp_row['webVideoUrl'],
            'trend_text': str(comp_row.get('text', '') if pd.notna(comp_row.get('text')) else '')[:60],
            'competitor_account': comp_row['author'],
            'competitor_momentum': comp_row['momentum_score'],
            'competitor_shares_h': comp_row.get('shares_per_hour', 0),
            'competitor_age_hours': comp_row.get('age_hours', 0),
            'you_also_posted': you_also_posted,
            'gap_type': gap_type,
            'hours_behind': hours_behind,
            'estimated_missed_revenue': round(potential_missed, 2),
            'market': comp_row.get('Market', 'Unknown'),
            'ai_category': comp_row.get('AI_CATEGORY', 'Unknown')
        })
    
    return pd.DataFrame(results)


def identify_competitor_patterns(df_historical: pd.DataFrame = None) -> Dict:
    """
    Identify patterns in competitor posting behavior.
    
    Returns insights like:
    - Preferred posting times
    - Hashtag preferences
    - Speed to market (how quickly they catch trends)
    """
    # This would analyze historical data
    # For now, return placeholder structure
    return {
        'avg_posting_hour_utc': None,
        'preferred_hashtags': [],
        'avg_trend_age_when_posted': None,
        'success_rate': None,
        'notes': 'Requires 7+ days of historical data for accurate patterns'
    }


def calculate_your_vs_competitor_metrics(
    df: pd.DataFrame,
    your_accounts: List[str] = YOUR_ACCOUNTS,
    competitor_accounts: List[str] = COMPETITOR_ACCOUNTS
) -> Dict:
    """Calculate head-to-head metrics."""
    
    df = df.copy()
    df['is_yours'] = df['author'].isin(your_accounts)
    df['is_competitor'] = df['author'].isin(competitor_accounts)
    
    your_posts = df[df['is_yours']]
    comp_posts = df[df['is_competitor']]
    
    return {
        'your_post_count': len(your_posts),
        'competitor_post_count': len(comp_posts),
        'your_avg_momentum': your_posts['momentum_score'].mean() if len(your_posts) > 0 else 0,
        'competitor_avg_momentum': comp_posts['momentum_score'].mean() if len(comp_posts) > 0 else 0,
        'your_total_momentum': your_posts['momentum_score'].sum() if len(your_posts) > 0 else 0,
        'competitor_total_momentum': comp_posts['momentum_score'].sum() if len(comp_posts) > 0 else 0,
        'your_spiking_count': len(your_posts[your_posts.get('acceleration_status', '').str.contains('SPIKING', na=False)]) if 'acceleration_status' in your_posts.columns else 0,
        'competitor_spiking_count': len(comp_posts[comp_posts.get('acceleration_status', '').str.contains('SPIKING', na=False)]) if 'acceleration_status' in comp_posts.columns else 0,
    }


# =============================================================================
# EXCEL OUTPUT GENERATION
# =============================================================================

def create_enhanced_excel(
    df_today: pd.DataFrame,
    df_yesterday: pd.DataFrame = None,
    df_2days_ago: pd.DataFrame = None,
    output_path: str = 'BUILD_TODAY_ENHANCED.xlsx',
    cache_path: str = None,
    dashboard_path: str = None,
    live_revenue_df: pd.DataFrame = None
) -> str:
    """
    Create v3.6.0 Enhanced Excel file with 10 tabs:
    1. DASHBOARD - Formula-driven KPI summary
    2. OPPORTUNITY_NOW - 13-column priority build list
    3. REVENUE_TRACKER - 20-column revenue tracking (carried forward)
    4. REVENUE_INSIGHTS - Auto-calculated breakdowns
    5. COMPETITOR_VIEW - 12-column combined competitor analysis
    6. PREDICTION_LOG - Model accuracy tracking
    7. DATA_FEED - 19-column enhanced MY_PERFORMANCE
    8. COMPETITOR_INTEL - 7-day deep competitor intelligence (9 sections)
    9. PAYMENTS - Day-by-day Pioneer Programme payments sorted by post date
    10. MONTHLY_REVENUE - Month-by-month revenue summary with totals
    """
    df_today = _ensure_calculated_metrics(df_today)

    # Filter to fresh content (72h) for enhanced analysis
    fresh_df = df_today[df_today['age_hours'] <= 72].copy() if 'age_hours' in df_today.columns else df_today.copy()

    # Calculate velocity predictions on fresh data
    df_with_predictions = calculate_velocity_predictions(fresh_df, df_yesterday, df_2days_ago)
    velocity_summary = create_velocity_summary(df_with_predictions, cache_path=cache_path)

    # Competitor analysis on full dataset
    competitor_gaps = analyze_competitor_gaps(df_today)
    h2h_metrics = calculate_your_vs_competitor_metrics(df_today)

    # Load existing revenue/prediction data
    existing_revenue = _load_existing_revenue(dashboard_path, live_revenue_df=live_revenue_df)
    existing_prediction_log = _load_existing_prediction_log(dashboard_path)

    # Build competitor intel (7-day deep analysis)
    try:
        from competitor_intel_patch import build_competitor_intel, build_competitor_intel_tab
        comp_cache_dir = os.path.dirname(cache_path) if cache_path else 'data'
        if not comp_cache_dir:
            comp_cache_dir = 'data'
        comp_intel = build_competitor_intel(df_today, comp_cache_dir)
        has_comp_intel = True
    except Exception as e:
        print(f"  [WARNING] Could not build competitor intel: {e}")
        comp_intel = None
        has_comp_intel = False

    # Style definitions
    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(bold=True, color='FFFFFF')
    cyan_fill = PatternFill('solid', fgColor='E0FFFF')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    wb = Workbook()

    # TAB 1: DASHBOARD
    ws = wb.active
    ws.title = 'DASHBOARD'
    _build_dashboard_tab(ws, header_fill, header_font)

    # TAB 2: OPPORTUNITY_NOW (13 columns)
    ws_opp = wb.create_sheet('OPPORTUNITY_NOW')
    _build_opportunity_now_tab(ws_opp, df_with_predictions, header_fill, header_font, thin_border)

    # TAB 3: REVENUE_TRACKER (19 columns)
    ws_rev = wb.create_sheet('REVENUE_TRACKER')
    _build_revenue_tracker_tab(ws_rev, existing_revenue, header_fill, header_font, thin_border)

    # TAB 4: REVENUE_INSIGHTS
    ws_ins = wb.create_sheet('REVENUE_INSIGHTS')
    _build_revenue_insights_tab(ws_ins, header_fill, header_font)

    # TAB 5: COMPETITOR_VIEW (12 columns)
    ws_comp = wb.create_sheet('COMPETITOR_VIEW')
    _build_competitor_view_tab(ws_comp, competitor_gaps, header_fill, header_font, thin_border)

    # TAB 6: PREDICTION_LOG (10 columns)
    ws_pred = wb.create_sheet('PREDICTION_LOG')
    _build_prediction_log_tab(ws_pred, df_with_predictions, df_yesterday,
                               existing_prediction_log, header_fill, header_font, thin_border)

    # TAB 7: DATA_FEED (19 columns)
    ws_feed = wb.create_sheet('DATA_FEED')
    _build_data_feed_tab(ws_feed, df_today, header_fill, header_font, thin_border, cyan_fill)

    # TAB 8: COMPETITOR_INTEL (7-day deep intelligence)
    if has_comp_intel and comp_intel is not None:
        ws_intel = wb.create_sheet('COMPETITOR_INTEL')
        build_competitor_intel_tab(ws_intel, comp_intel, header_fill, header_font, thin_border)

    # TAB 9: PAYMENTS (day-by-day Pioneer Programme payments)
    ws_pay = wb.create_sheet('PAYMENTS')
    _build_payments_tab(ws_pay, existing_revenue, header_fill, header_font, thin_border)

    # TAB 10: MONTHLY_REVENUE (month-by-month summary)
    ws_monthly = wb.create_sheet('MONTHLY_REVENUE')
    _build_monthly_revenue_tab(ws_monthly, existing_revenue, header_fill, header_font, thin_border)

    wb.save(output_path)
    return output_path


# =============================================================================
# v3.6.0 TAB BUILDERS
# =============================================================================

def _build_dashboard_tab(ws, header_fill, header_font):
    section_fill = PatternFill('solid', fgColor='E8F4FD')  # Light blue for section headers
    kpi_fill = PatternFill('solid', fgColor='F5F5F5')  # Light grey for KPI cells
    red_kpi = PatternFill('solid', fgColor='FFE0E0')  # Light red for urgent KPIs
    green_kpi = PatternFill('solid', fgColor='E0FFE0')  # Light green for positive KPIs
    yellow_kpi = PatternFill('solid', fgColor='FFFDE0')  # Light yellow for seasonal
    
    ws['A1'] = '\U0001f4ca TIKTOK TREND SYSTEM \u2014 DAILY DASHBOARD'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].fill = PatternFill('solid', fgColor='1F4E78')
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws.merge_cells('A1:L1')
    ws['A2'] = '=TEXT(NOW(),"dddd, dd mmmm yyyy \u2014 HH:MM")'
    ws['A2'].font = Font(italic=True, color='666666')
    ws.merge_cells('A2:F2')

    # REVENUE section
    ws['A4'] = '\U0001f4b0 REVENUE'
    ws['A4'].font = Font(bold=True, size=12, color='1F4E78')
    for col in range(1, 13):
        ws.cell(row=4, column=col).fill = section_fill
    ws.merge_cells('A4:L4')
    for i, lbl in enumerate(['Total Revenue', '', 'Templates at Cap', '', 'Hit Rate', '', 'Avg Rev/Template', '', 'Outstanding'], 1):
        ws.cell(row=5, column=i, value=lbl).font = Font(bold=True, size=9, color='666666')
    ws['A6'] = '=SUMPRODUCT(REVENUE_TRACKER!D2:D1000)'
    ws['A6'].number_format = '$#,##0'
    ws['A6'].font = Font(bold=True, size=16)
    ws['A6'].fill = green_kpi
    ws['C6'] = '=COUNTIF(REVENUE_TRACKER!D2:D1000,">=2500")'
    ws['C6'].font = Font(bold=True, size=16)
    ws['E6'] = '=IFERROR(COUNTIF(REVENUE_TRACKER!D2:D1000,">0")/COUNTA(REVENUE_TRACKER!A2:A1000),0)'
    ws['E6'].number_format = '0%'
    ws['E6'].font = Font(bold=True, size=16)
    ws['G6'] = '=IFERROR(AVERAGEIF(REVENUE_TRACKER!D2:D1000,">0"),0)'
    ws['G6'].number_format = '$#,##0'
    ws['G6'].font = Font(bold=True, size=16)
    ws['I6'] = '=SUMPRODUCT(REVENUE_TRACKER!E2:E1000)-SUMPRODUCT(REVENUE_TRACKER!D2:D1000)'
    ws['I6'].number_format = '$#,##0'
    ws['I6'].font = Font(bold=True, size=16)

    # TODAY section
    ws['A9'] = '\U0001f3af TODAY'
    ws['A9'].font = Font(bold=True, size=12, color='1F4E78')
    for col in range(1, 13):
        ws.cell(row=9, column=col).fill = section_fill
    ws.merge_cells('A9:L9')
    for i, lbl in enumerate(['Build NOW', '', 'Build Today', '', 'URGENT', '', 'SPIKING', '', 'Your Posts'], 1):
        ws.cell(row=10, column=i, value=lbl).font = Font(bold=True, size=9, color='666666')
    ws['A11'] = '=COUNTIF(OPPORTUNITY_NOW!B2:B100,"*BUILD_IMMEDIATELY*")'
    ws['A11'].font = Font(bold=True, size=16, color='FF0000')
    ws['A11'].fill = red_kpi
    ws['C11'] = '=COUNTIF(OPPORTUNITY_NOW!B2:B100,"*BUILD_TODAY*")'
    ws['C11'].font = Font(bold=True, size=16, color='FF8C00')
    ws['C11'].fill = PatternFill('solid', fgColor='FFF3E0')
    ws['E11'] = '=COUNTIF(DATA_FEED!L2:L5000,"*URGENT*")'
    ws['E11'].font = Font(bold=True, size=16)
    ws['E11'].fill = kpi_fill
    ws['G11'] = '=COUNTIF(DATA_FEED!F2:F5000,"*SPIKING*")'
    ws['G11'].font = Font(bold=True, size=16)
    ws['G11'].fill = kpi_fill
    ws['I11'] = '=COUNTA(DATA_FEED!B2:B5000)'
    ws['I11'].font = Font(bold=True, size=16)
    ws['I11'].fill = kpi_fill

    # COMPETITOR & MODEL section
    ws['A14'] = '\u2694\ufe0f COMPETITOR & MODEL'
    ws['A14'].font = Font(bold=True, size=12, color='1F4E78')
    for col in range(1, 13):
        ws.cell(row=14, column=col).fill = section_fill
    ws.merge_cells('A14:L14')
    for i, lbl in enumerate(['You Missed', '', 'Competitor Posts', '', 'Prediction Accuracy', '', 'Model Bias'], 1):
        ws.cell(row=15, column=i, value=lbl).font = Font(bold=True, size=9, color='666666')
    ws['A16'] = '=COUNTIF(COMPETITOR_VIEW!H2:H500,"MISSED_BY_YOU")'
    ws['A16'].font = Font(bold=True, size=16, color='FF0000')
    ws['A16'].fill = red_kpi
    ws['C16'] = '=COUNTA(COMPETITOR_VIEW!A2:A500)'
    ws['C16'].font = Font(bold=True, size=16)
    ws['C16'].fill = kpi_fill
    ws['E16'] = '=IFERROR(INDEX(PREDICTION_LOG!C2:C100,MATCH(9.99E+307,PREDICTION_LOG!C2:C100)),0)'
    ws['E16'].number_format = '0%'
    ws['E16'].font = Font(bold=True, size=16)
    ws['E16'].fill = kpi_fill
    ws['G16'] = '=IFERROR(INDEX(PREDICTION_LOG!D2:D100,MATCH(9.99E+307,PREDICTION_LOG!D2:D100)),"N/A")'
    ws['G16'].font = Font(bold=True, size=16)
    ws['G16'].fill = kpi_fill

    # SEASONAL
    ws['A19'] = '\U0001f4c5 SEASONAL ALERTS'
    ws['A19'].font = Font(bold=True, size=12, color='1F4E78')
    for col in range(1, 13):
        ws.cell(row=19, column=col).fill = yellow_kpi
    ws.merge_cells('A19:L19')
    seasonal_end_row = 20  # Track where seasonal section ends
    try:
        from seasonal_calendar import get_seasonal_alerts
        from datetime import date
        alerts = get_seasonal_alerts(date.today())
        if alerts:
            for i, a in enumerate(alerts):
                cell = ws.cell(row=20+i, column=1, value=f"{a.get('emoji','')} {a.get('event','')} \u2014 {a.get('timing','')}")
                # Color based on priority
                pri = a.get('priority_level', '')
                if pri in ('today', 'urgent'):
                    cell.font = Font(bold=True, color='CC0000')
                elif pri == 'review':
                    cell.font = Font(color='996600')
                elif pri == 'high':
                    cell.font = Font(bold=True, color='006600')
                elif pri == 'window':
                    cell.font = Font(italic=True, color='1F4E78')
                else:
                    cell.font = Font(color='333333')
            seasonal_end_row = 20 + len(alerts)
        else:
            ws['A20'] = 'No seasonal events coming up'
            seasonal_end_row = 21
    except Exception:
        ws['A20'] = 'Seasonal calendar not available'
        seasonal_end_row = 21

    # HOW TO USE (dynamically positioned after seasonal)
    how_to_row = seasonal_end_row + 1
    ws.cell(row=how_to_row, column=1, value='\U0001f4cb HOW TO USE THIS DASHBOARD')
    ws.cell(row=how_to_row, column=1).font = Font(bold=True, size=12, color='1F4E78')
    for col in range(1, 13):
        ws.cell(row=how_to_row, column=col).fill = section_fill
    ws.merge_cells(start_row=how_to_row, start_column=1, end_row=how_to_row, end_column=12)
    instructions = [
        '1. Open OPPORTUNITY_NOW tab > build the red items first',
        '2. Check COMPETITOR_VIEW > look for MISSED_BY_YOU with high momentum',
        '3. After building templates, fill in REVENUE_TRACKER with your template links',
        '4. When revenue comes in, update the Estimated and Received columns',
        '5. REVENUE_INSIGHTS auto-calculates which signals predict revenue best',
        '6. PREDICTION_LOG tracks model accuracy > check for tuning suggestions',
    ]
    for i, txt in enumerate(instructions):
        ws.cell(row=how_to_row+1+i, column=1, value=txt).font = Font(size=10)
    for col in range(1, 13):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col)].width = 16


def _build_opportunity_now_tab(ws, df_pred, header_fill, header_font, thin_border):
    headers = ['Priority', 'Build Priority', 'Time Zone', 'Time Remaining',
               'Trend', 'Creator', 'Momentum', 'Opportunity Score', 'Age',
               'Market', 'Seasonal', 'Previously Built', 'URL']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center')

    # Filter to actionable
    actionable = df_pred[
        (df_pred['action_window'].str.contains('ACT NOW|6-12H', na=False)) &
        (df_pred['age_hours'] <= 48) & (df_pred['momentum_score'] >= 500)
    ].copy()
    
    # Fallback 1: Expand to include 12-24H and MONITOR with decent momentum
    if len(actionable) == 0:
        actionable = df_pred[
            (df_pred['action_window'].str.contains('ACT NOW|6-12H|12-24H|MONITOR', na=False)) &
            (df_pred['age_hours'] <= 60) & (df_pred['momentum_score'] >= 300)
        ].nlargest(20, 'momentum_score').copy()
    
    # Fallback 2: If velocity data is missing (all PEAKED/TOO LATE), use pure momentum
    if len(actionable) == 0:
        actionable = df_pred[
            (df_pred['age_hours'] <= 60) & (df_pred['momentum_score'] >= 300)
        ].nlargest(20, 'momentum_score').copy()
    
    # Fallback 3: Just show top 20 by momentum regardless
    if len(actionable) == 0:
        actionable = df_pred[
            df_pred['age_hours'] <= 72
        ].nlargest(20, 'momentum_score').copy()

    # Exclude tracked accounts
    if 'author' in actionable.columns and len(actionable) > 0:
        all_tracked = [a.lower() for a in YOUR_ACCOUNTS + COMPETITOR_ACCOUNTS]
        actionable = actionable[~actionable['author'].str.lower().isin(all_tracked)]

    # Opportunity score
    if len(actionable) > 0:
        vel_fill = actionable['velocity'].fillna(0)
        age_factor = (72 - actionable['age_hours'].clip(0, 72)) / 72
        actionable['opportunity_score'] = (
            actionable['momentum_score'] * 0.4 + vel_fill.clip(lower=0) * 2 +
            actionable['predicted_24h'].fillna(actionable['momentum_score']) * 0.3 +
            age_factor * 1000
        ).astype(int)
        actionable = actionable.nlargest(20, 'opportunity_score')

    now_hour = datetime.utcnow().hour
    is_prime = 8 <= now_hour <= 22
    tz_label = '\U0001f7e2 PRIME' if is_prime else 'OFF_PEAK'
    seasonal_text = ''
    try:
        from seasonal_calendar import get_seasonal_alerts
        from datetime import date
        alerts = get_seasonal_alerts(date.today())
        if alerts: seasonal_text = alerts[0].get('event', '')
    except Exception: pass

    for ri, (_, row) in enumerate(actionable.iterrows(), 2):
        aw = str(row.get('action_window', ''))
        is_act_now = 'ACT NOW' in aw
        hours_left = max(0, 72 - row.get('age_hours', 0))
        time_rem = f"{hours_left:.0f}h of prime time left" if is_prime else f"{hours_left:.0f}h remaining"
        vals = [
            ri - 1,
            '\U0001f534 BUILD_IMMEDIATELY' if is_act_now else '\U0001f7e0 BUILD_TODAY',
            tz_label, time_rem,
            str(row.get('text', ''))[:60] if pd.notna(row.get('text')) else '',
            str(row.get('author', ''))[:20] if pd.notna(row.get('author')) else '',
            int(row.get('momentum_score', 0)),
            int(row.get('opportunity_score', 0)),
            f"{row.get('age_hours', 0):.1f}h",
            str(row.get('Market', '')) if pd.notna(row.get('Market')) else '',
            seasonal_text, '',
            str(row.get('webVideoUrl', '')),
        ]
        row_color = 'FF6B6B' if is_act_now else 'FFE4B5'
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=_sanitize_cell(val))
            c.border = thin_border
            if ci <= 12:
                c.fill = PatternFill('solid', fgColor=row_color)
                if is_act_now: c.font = Font(bold=True)
        url_cell = ws.cell(row=ri, column=13)
        url_val = str(row.get('webVideoUrl', ''))
        if url_val.startswith('http'):
            url_cell.hyperlink = url_val
            url_cell.font = Font(color='0000FF', underline='single')

    if len(actionable) == 0:
        ws['A2'] = 'No immediate opportunities - check DASHBOARD for monitoring items'
        ws['A2'].font = Font(italic=True, color='666666')

    ws.freeze_panes = 'A2'
    for col, w in [('A',8),('B',22),('C',12),('D',25),('E',50),('F',18),('G',12),('H',16),('I',10),('J',15),('K',18),('L',15),('M',50)]:
        ws.column_dimensions[col].width = w


def _build_revenue_tracker_tab(ws, existing_revenue, header_fill, header_font, thin_border):
    headers = ['TikTok URL', 'Account', 'Template Link', 'Received ($)',
               'Estimated ($)', 'US & EU3 Installs', 'ROW Installs',
               'Total Installs', 'Rev/Install', 'At Cap?', 'Trend Description',
               'Momentum at Detection', 'Trigger Level', 'Action Window',
               'Market', 'AI Category', 'Age at Detection', 'Date First Seen', 'Notes',
               'Post Date']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center')

    max_data_row = 1
    if existing_revenue is not None and len(existing_revenue) > 0:
        from revenue_model import extract_post_date_str
        for ri, (_, row) in enumerate(existing_revenue.iterrows(), 2):
            for ci in range(1, 20):
                col_name = headers[ci-1] if ci <= len(headers) else ''
                val = row.get(col_name, row.iloc[ci-1] if ci-1 < len(row) else '')
                if pd.isna(val): val = ''
                ws.cell(row=ri, column=ci, value=_sanitize_cell(val)).border = thin_border
            # Column 20: Post Date (derived from video URL)
            url_val = str(row.get('TikTok URL', row.get('url', '')))
            post_date = extract_post_date_str(url_val, '%Y-%m-%d %H:%M')
            ws.cell(row=ri, column=20, value=post_date).border = thin_border
            # Formulas for calculated columns
            ws.cell(row=ri, column=8, value=f'=F{ri}+G{ri}')
            ws.cell(row=ri, column=9, value=f'=IFERROR(E{ri}/H{ri},0)')
            ws.cell(row=ri, column=9).number_format = '$#,##0.00'
            ws.cell(row=ri, column=10, value=f'=IF(E{ri}>=2500,"\u2705 CAP","")')
            max_data_row = ri

    # Add formulas for empty rows (for future user input)
    input_fill = PatternFill('solid', fgColor='FFFFE0')
    for ri in range(max_data_row + 1, max_data_row + 51):
        ws.cell(row=ri, column=8, value=f'=F{ri}+G{ri}')
        ws.cell(row=ri, column=9, value=f'=IFERROR(E{ri}/H{ri},0)')
        ws.cell(row=ri, column=9).number_format = '$#,##0.00'
        ws.cell(row=ri, column=10, value=f'=IF(E{ri}>=2500,"\u2705 CAP","")')
        for ci in [3,4,5,6,7,19]:
            ws.cell(row=ri, column=ci).fill = input_fill

    ws.freeze_panes = 'A2'
    for col, w in [('A',50),('B',20),('C',40),('D',12),('E',12),('F',15),('G',12),('H',12),('I',12),('J',10),('K',40),('R',14),('T',18)]:
        ws.column_dimensions[col].width = w


def _build_payments_tab(ws, existing_revenue, header_fill, header_font, thin_border):
    """
    PAYMENTS tab — Day-by-day Pioneer Programme payments.
    Sorted by post date (derived from video ID), with daily subtotals.
    Clean, easy-to-read view of all template earnings.
    """
    from revenue_model import extract_post_date_str, extract_post_date
    
    headers = ['Post Date', 'TikTok URL', 'Received ($)', 'Estimated ($)',
               'US & EU3 Installs', 'ROW Installs', 'Total Installs',
               'Rev/Install', 'At Cap?', 'Status']
    
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center')
    
    if existing_revenue is None or len(existing_revenue) == 0:
        ws.cell(row=2, column=1, value='No revenue data available')
        ws.freeze_panes = 'A2'
        return
    
    # Build list of entries with post dates
    entries = []
    for _, row in existing_revenue.iterrows():
        url = str(row.get('TikTok URL', row.get('url', '')))
        if not url or not url.startswith('http'):
            continue
        post_dt = extract_post_date(url)
        post_date_str = post_dt.strftime('%Y-%m-%d') if post_dt else ''
        received = float(row.get('Received ($)', row.get('received', 0))) if pd.notna(row.get('Received ($)', row.get('received', 0))) else 0
        estimated = float(row.get('Estimated ($)', row.get('estimated', 0))) if pd.notna(row.get('Estimated ($)', row.get('estimated', 0))) else 0
        us_inst = float(row.get('US & EU3 Installs', row.get('us_installs', 0))) if pd.notna(row.get('US & EU3 Installs', row.get('us_installs', 0))) else 0
        row_inst = float(row.get('ROW Installs', row.get('row_installs', 0))) if pd.notna(row.get('ROW Installs', row.get('row_installs', 0))) else 0
        total_inst = us_inst + row_inst
        rev_per = estimated / total_inst if total_inst > 0 else 0
        at_cap = '✅ CAP' if received >= 2500 else ''
        
        # Status
        if received >= 2500:
            status = '🟢 Capped'
        elif received > 0:
            status = '🟡 Earning'
        elif estimated > 0:
            status = '🟠 Pending'
        elif total_inst > 0:
            status = '⚪ Installs only'
        else:
            status = '⭕ No activity'
        
        entries.append({
            'post_date_str': post_date_str,
            'post_dt': post_dt,
            'url': url,
            'received': received,
            'estimated': estimated,
            'us_installs': int(us_inst),
            'row_installs': int(row_inst),
            'total_installs': int(total_inst),
            'rev_per': rev_per,
            'at_cap': at_cap,
            'status': status,
        })
    
    # Sort by post date (newest first)
    from datetime import datetime as _dt_cls, timezone as _tz
    _min_dt = _dt_cls(2000, 1, 1, tzinfo=_tz.utc)
    entries.sort(key=lambda x: x['post_dt'] or _min_dt, reverse=True)
    
    # Color definitions
    cap_fill = PatternFill('solid', fgColor='C6EFCE')     # Green for capped
    earning_fill = PatternFill('solid', fgColor='FFFFE0')  # Light yellow for earning
    pending_fill = PatternFill('solid', fgColor='FFE4B5')  # Orange for pending
    
    ri = 2
    current_date = None
    daily_received = 0
    daily_estimated = 0
    daily_start_row = 2
    grand_received = 0
    grand_estimated = 0
    grand_installs = 0
    
    for entry in entries:
        date_str = entry['post_date_str']
        
        # Insert daily subtotal when date changes
        if current_date is not None and date_str != current_date and daily_received + daily_estimated > 0:
            # Subtotal row
            sub_fill = PatternFill('solid', fgColor='D9E1F2')
            ws.cell(row=ri, column=1, value=f'  Subtotal: {current_date}').font = Font(bold=True, size=10)
            ws.cell(row=ri, column=3, value=daily_received).number_format = '$#,##0.00'
            ws.cell(row=ri, column=4, value=daily_estimated).number_format = '$#,##0.00'
            for ci in range(1, 11):
                ws.cell(row=ri, column=ci).fill = sub_fill
                ws.cell(row=ri, column=ci).border = thin_border
            ws.cell(row=ri, column=3).font = Font(bold=True)
            ws.cell(row=ri, column=4).font = Font(bold=True)
            ri += 1
            daily_received = 0
            daily_estimated = 0
        
        current_date = date_str
        daily_received += entry['received']
        daily_estimated += entry['estimated']
        grand_received += entry['received']
        grand_estimated += entry['estimated']
        grand_installs += entry['total_installs']
        
        # Data row
        vals = [entry['post_date_str'], entry['url'], entry['received'],
                entry['estimated'], entry['us_installs'], entry['row_installs'],
                entry['total_installs'], entry['rev_per'], entry['at_cap'], entry['status']]
        
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = thin_border
            if ci == 3 or ci == 4:
                c.number_format = '$#,##0.00'
            elif ci == 8:
                c.number_format = '$#,##0.00'
        
        # Row coloring based on status
        if entry['received'] >= 2500:
            for ci in range(1, 11):
                ws.cell(row=ri, column=ci).fill = cap_fill
        elif entry['received'] > 0:
            for ci in range(1, 11):
                ws.cell(row=ri, column=ci).fill = earning_fill
        elif entry['estimated'] > 0:
            for ci in range(1, 11):
                ws.cell(row=ri, column=ci).fill = pending_fill
        
        # URL as hyperlink
        try:
            ws.cell(row=ri, column=2).hyperlink = entry['url']
            ws.cell(row=ri, column=2).font = Font(color='0000FF', underline='single')
        except Exception:
            pass
        
        ri += 1
    
    # Final daily subtotal
    if current_date is not None and daily_received + daily_estimated > 0:
        sub_fill = PatternFill('solid', fgColor='D9E1F2')
        ws.cell(row=ri, column=1, value=f'  Subtotal: {current_date}').font = Font(bold=True, size=10)
        ws.cell(row=ri, column=3, value=daily_received).number_format = '$#,##0.00'
        ws.cell(row=ri, column=4, value=daily_estimated).number_format = '$#,##0.00'
        for ci in range(1, 11):
            ws.cell(row=ri, column=ci).fill = sub_fill
            ws.cell(row=ri, column=ci).border = thin_border
        ri += 1
    
    # Grand total
    ri += 1
    total_fill = PatternFill('solid', fgColor='1F4E78')
    total_font = Font(bold=True, color='FFFFFF', size=11)
    ws.cell(row=ri, column=1, value='GRAND TOTAL').fill = total_fill
    ws.cell(row=ri, column=1).font = total_font
    ws.cell(row=ri, column=3, value=grand_received).fill = total_fill
    ws.cell(row=ri, column=3).font = total_font
    ws.cell(row=ri, column=3).number_format = '$#,##0.00'
    ws.cell(row=ri, column=4, value=grand_estimated).fill = total_fill
    ws.cell(row=ri, column=4).font = total_font
    ws.cell(row=ri, column=4).number_format = '$#,##0.00'
    ws.cell(row=ri, column=7, value=grand_installs).fill = total_fill
    ws.cell(row=ri, column=7).font = total_font
    for ci in [2, 5, 6, 8, 9, 10]:
        ws.cell(row=ri, column=ci).fill = total_fill
    
    ws.freeze_panes = 'A2'
    for col, w in [('A', 14), ('B', 55), ('C', 14), ('D', 14), ('E', 16), ('F', 14), ('G', 14), ('H', 12), ('I', 10), ('J', 16)]:
        ws.column_dimensions[col].width = w


def _build_monthly_revenue_tab(ws, existing_revenue, header_fill, header_font, thin_border):
    """
    MONTHLY_REVENUE tab — Month-by-month revenue summary.
    Shows templates posted, revenue earned, installs, cap rate per month.
    """
    from revenue_model import extract_post_month, extract_post_date
    
    if existing_revenue is None or len(existing_revenue) == 0:
        ws.cell(row=1, column=1, value='No revenue data available')
        return
    
    # Build monthly aggregations
    monthly = {}
    for _, row in existing_revenue.iterrows():
        url = str(row.get('TikTok URL', row.get('url', '')))
        if not url or not url.startswith('http'):
            continue
        month = extract_post_month(url)
        if not month:
            continue
        
        received = float(row.get('Received ($)', row.get('received', 0))) if pd.notna(row.get('Received ($)', row.get('received', 0))) else 0
        estimated = float(row.get('Estimated ($)', row.get('estimated', 0))) if pd.notna(row.get('Estimated ($)', row.get('estimated', 0))) else 0
        us_inst = float(row.get('US & EU3 Installs', row.get('us_installs', 0))) if pd.notna(row.get('US & EU3 Installs', row.get('us_installs', 0))) else 0
        row_inst = float(row.get('ROW Installs', row.get('row_installs', 0))) if pd.notna(row.get('ROW Installs', row.get('row_installs', 0))) else 0
        
        if month not in monthly:
            monthly[month] = {
                'templates': 0, 'received': 0, 'estimated': 0,
                'us_installs': 0, 'row_installs': 0, 'total_installs': 0,
                'at_cap': 0, 'with_revenue': 0,
            }
        m = monthly[month]
        m['templates'] += 1
        m['received'] += received
        m['estimated'] += estimated
        m['us_installs'] += us_inst
        m['row_installs'] += row_inst
        m['total_installs'] += us_inst + row_inst
        if received >= 2500:
            m['at_cap'] += 1
        if received > 0:
            m['with_revenue'] += 1
    
    # Headers
    headers = ['Month', 'Templates', 'Received ($)', 'Estimated ($)',
               'US & EU3 Installs', 'ROW Installs', 'Total Installs',
               'At Cap', 'With Revenue', 'Cap Rate (%)', 'Avg Rev/Template',
               'Rev/Install']
    
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center')
    
    # Data rows (sorted by month)
    ri = 2
    total_templates = 0
    total_received = 0
    total_estimated = 0
    total_installs = 0
    total_us = 0
    total_row = 0
    total_cap = 0
    total_with_rev = 0
    
    for month in sorted(monthly.keys()):
        m = monthly[month]
        cap_rate = (m['at_cap'] / m['templates'] * 100) if m['templates'] > 0 else 0
        avg_rev = m['received'] / m['templates'] if m['templates'] > 0 else 0
        rev_per_install = m['received'] / m['total_installs'] if m['total_installs'] > 0 else 0
        
        vals = [month, m['templates'], m['received'], m['estimated'],
                int(m['us_installs']), int(m['row_installs']), int(m['total_installs']),
                m['at_cap'], m['with_revenue'], round(cap_rate, 1), round(avg_rev, 2),
                round(rev_per_install, 2)]
        
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = thin_border
            if ci in [3, 4, 11]:
                c.number_format = '$#,##0.00'
            elif ci == 12:
                c.number_format = '$#,##0.00'
            elif ci == 10:
                c.number_format = '0.0%'
                c.value = cap_rate / 100  # Store as decimal for Excel
        
        # Color highlight months with high revenue
        if m['received'] >= 10000:
            month_fill = PatternFill('solid', fgColor='C6EFCE')  # Green
        elif m['received'] >= 1000:
            month_fill = PatternFill('solid', fgColor='FFFFE0')  # Light yellow
        else:
            month_fill = None
        
        if month_fill:
            for ci in range(1, 13):
                ws.cell(row=ri, column=ci).fill = month_fill
        
        total_templates += m['templates']
        total_received += m['received']
        total_estimated += m['estimated']
        total_us += m['us_installs']
        total_row += m['row_installs']
        total_installs += m['total_installs']
        total_cap += m['at_cap']
        total_with_rev += m['with_revenue']
        ri += 1
    
    # Grand total row
    ri += 1
    total_fill = PatternFill('solid', fgColor='1F4E78')
    total_font = Font(bold=True, color='FFFFFF', size=11)
    
    totals = ['TOTAL', total_templates, total_received, total_estimated,
              int(total_us), int(total_row), int(total_installs),
              total_cap, total_with_rev,
              total_cap / total_templates if total_templates > 0 else 0,
              total_received / total_templates if total_templates > 0 else 0,
              total_received / total_installs if total_installs > 0 else 0]
    
    for ci, val in enumerate(totals, 1):
        c = ws.cell(row=ri, column=ci, value=val)
        c.fill = total_fill
        c.font = total_font
        if ci in [3, 4, 11]:
            c.number_format = '$#,##0.00'
        elif ci == 12:
            c.number_format = '$#,##0.00'
        elif ci == 10:
            c.number_format = '0.0%'
    
    # Month-over-month growth section
    ri += 3
    ws.cell(row=ri, column=1, value='MONTH-OVER-MONTH GROWTH').font = Font(bold=True, size=12, color='1F4E78')
    ri += 1
    growth_headers = ['Month', 'Revenue', 'Change ($)', 'Change (%)', 'Templates', 'Pace']
    for ci, h in enumerate(growth_headers, 1):
        c = ws.cell(row=ri, column=ci, value=h)
        c.fill = PatternFill('solid', fgColor='D9E1F2')
        c.font = Font(bold=True)
    
    ri += 1
    sorted_months = sorted(monthly.keys())
    prev_rev = 0
    for month in sorted_months:
        m = monthly[month]
        change = m['received'] - prev_rev
        change_pct = (change / prev_rev * 100) if prev_rev > 0 else 0
        pace = '📈 Growing' if change > 0 else ('📉 Declining' if change < 0 else '➡️ Flat')
        
        ws.cell(row=ri, column=1, value=month).border = thin_border
        ws.cell(row=ri, column=2, value=m['received']).border = thin_border
        ws.cell(row=ri, column=2).number_format = '$#,##0.00'
        ws.cell(row=ri, column=3, value=change).border = thin_border
        ws.cell(row=ri, column=3).number_format = '$#,##0.00'
        if prev_rev > 0:
            ws.cell(row=ri, column=4, value=change_pct / 100).border = thin_border
            ws.cell(row=ri, column=4).number_format = '0.0%'
        else:
            ws.cell(row=ri, column=4, value='N/A').border = thin_border
        ws.cell(row=ri, column=5, value=m['templates']).border = thin_border
        ws.cell(row=ri, column=6, value=pace).border = thin_border
        
        prev_rev = m['received']
        ri += 1
    
    ws.freeze_panes = 'A2'
    for col, w in [('A', 14), ('B', 12), ('C', 14), ('D', 14), ('E', 16), ('F', 14), ('G', 14), ('H', 10), ('I', 14), ('J', 12), ('K', 16), ('L', 12)]:
        ws.column_dimensions[col].width = w


def _build_revenue_insights_tab(ws, header_fill, header_font):
    """
    REVENUE_INSIGHTS tab — matches user's target structure exactly.
    4 sections: By Trigger Level (6 cols), By Market (5 cols), By AI Category (5 cols), Install Economics (2 cols).
    All formulas reference REVENUE_TRACKER tab.
    """
    section_font = Font(bold=True, size=12, color='1F4E78')
    table_header_fill = PatternFill('solid', fgColor='1F4E78')
    table_header_font = Font(bold=True, size=11, color='FFFFFF')

    # === ROW 1: Title ===
    ws['A1'] = '\U0001f4b0 REVENUE INSIGHTS \u2014 Auto-Calculated'
    ws['A1'].font = Font(bold=True, size=16, color='1F4E78')
    ws.merge_cells('A1:F1')

    # === SECTION 1: REVENUE BY TRIGGER LEVEL (rows 3-8) ===
    ws['A3'] = 'REVENUE BY TRIGGER LEVEL'
    ws['A3'].font = section_font

    # Headers row 4
    for ci, h in enumerate(['Trigger Level', 'Templates', 'Total Revenue', 'Avg Revenue', 'Hit Rate', 'Cap Rate', 'Best Template'], 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.fill = table_header_fill; c.font = table_header_font

    # Data rows 5-8
    for ri, trig in enumerate(['URGENT', 'HIGH', 'WATCH', 'NONE'], 5):
        ws.cell(row=ri, column=1, value=trig)
        ws.cell(row=ri, column=2, value=f'=COUNTIF(REVENUE_TRACKER!M:M,"*{trig}*")')
        ws.cell(row=ri, column=3, value=f'=SUMPRODUCT((ISNUMBER(SEARCH("{trig}",REVENUE_TRACKER!M$2:M$1000)))*REVENUE_TRACKER!D$2:D$1000)')
        ws.cell(row=ri, column=3).number_format = '$#,##0'
        ws.cell(row=ri, column=4, value=f'=IFERROR(C{ri}/B{ri},0)')
        ws.cell(row=ri, column=4).number_format = '$#,##0'
        ws.cell(row=ri, column=5, value=f'=IFERROR(SUMPRODUCT((ISNUMBER(SEARCH("{trig}",REVENUE_TRACKER!M$2:M$1000)))*(REVENUE_TRACKER!D$2:D$1000>0))/B{ri},0)')
        ws.cell(row=ri, column=5).number_format = '0%'
        ws.cell(row=ri, column=6, value=f'=IFERROR(SUMPRODUCT((ISNUMBER(SEARCH("{trig}",REVENUE_TRACKER!M$2:M$1000)))*(REVENUE_TRACKER!D$2:D$1000>=2500))/B{ri},0)')
        ws.cell(row=ri, column=6).number_format = '0%'
        ws.cell(row=ri, column=7, value=f'=IFERROR(MAXIFS(REVENUE_TRACKER!D$2:D$1000,REVENUE_TRACKER!M$2:M$1000,"*{trig}*"),0)')
        ws.cell(row=ri, column=7).number_format = '$#,##0'

    # === SECTION 2: REVENUE BY MARKET (rows 11-15) ===
    ws['A11'] = 'REVENUE BY MARKET'
    ws['A11'].font = section_font

    # Headers row 12
    for ci, h in enumerate(['Market', 'Templates', 'Total Revenue', 'Avg Revenue', 'Cap Rate', 'Best Template'], 1):
        c = ws.cell(row=12, column=ci, value=h)
        c.fill = table_header_fill; c.font = table_header_font

    # Data rows 13-15
    for ri, (mkt, search_term) in enumerate([('BOTH', 'BOTH'), ('US ONLY', 'US ONLY'), ('UK ONLY', 'UK ONLY')], 13):
        ws.cell(row=ri, column=1, value=mkt)
        ws.cell(row=ri, column=2, value=f'=COUNTIF(REVENUE_TRACKER!O:O,"*{search_term}*")')
        ws.cell(row=ri, column=3, value=f'=SUMPRODUCT((ISNUMBER(SEARCH("{search_term}",REVENUE_TRACKER!O$2:O$1000)))*REVENUE_TRACKER!D$2:D$1000)')
        ws.cell(row=ri, column=3).number_format = '$#,##0'
        ws.cell(row=ri, column=4, value=f'=IFERROR(C{ri}/B{ri},0)')
        ws.cell(row=ri, column=4).number_format = '$#,##0'
        ws.cell(row=ri, column=5, value=f'=IFERROR(SUMPRODUCT((ISNUMBER(SEARCH("{search_term}",REVENUE_TRACKER!O$2:O$1000)))*(REVENUE_TRACKER!D$2:D$1000>=2500))/B{ri},0)')
        ws.cell(row=ri, column=5).number_format = '0%'
        ws.cell(row=ri, column=6, value=f'=IFERROR(MAXIFS(REVENUE_TRACKER!D$2:D$1000,REVENUE_TRACKER!O$2:O$1000,"*{search_term}*"),0)')
        ws.cell(row=ri, column=6).number_format = '$#,##0'

    # === SECTION 3: REVENUE BY AI CATEGORY (rows 18-21) ===
    ws['A18'] = 'REVENUE BY AI CATEGORY'
    ws['A18'].font = section_font

    # Headers row 19
    for ci, h in enumerate(['Category', 'Templates', 'Total Revenue', 'Avg Revenue', 'Cap Rate', 'Best Template'], 1):
        c = ws.cell(row=19, column=ci, value=h)
        c.fill = table_header_fill; c.font = table_header_font

    # Data rows 20-21 (exact match uses = not SEARCH for AI category)
    ws.cell(row=20, column=1, value='AI')
    ws.cell(row=20, column=2, value='=COUNTIF(REVENUE_TRACKER!P:P,"*AI")')
    ws.cell(row=20, column=3, value='=SUMPRODUCT((REVENUE_TRACKER!P$2:P$1000="AI")*REVENUE_TRACKER!D$2:D$1000)')
    ws.cell(row=20, column=3).number_format = '$#,##0'
    ws.cell(row=20, column=4, value='=IFERROR(C20/B20,0)')
    ws.cell(row=20, column=4).number_format = '$#,##0'
    ws.cell(row=20, column=5, value='=IFERROR(SUMPRODUCT((REVENUE_TRACKER!P$2:P$1000="AI")*(REVENUE_TRACKER!D$2:D$1000>=2500))/B20,0)')
    ws.cell(row=20, column=5).number_format = '0%'
    ws.cell(row=20, column=6, value='=IFERROR(MAXIFS(REVENUE_TRACKER!D$2:D$1000,REVENUE_TRACKER!P$2:P$1000,"AI"),0)')
    ws.cell(row=20, column=6).number_format = '$#,##0'

    ws.cell(row=21, column=1, value='NON-AI')
    ws.cell(row=21, column=2, value='=COUNTIF(REVENUE_TRACKER!P:P,"*NON-AI")')
    ws.cell(row=21, column=3, value='=SUMPRODUCT((REVENUE_TRACKER!P$2:P$1000="NON-AI")*REVENUE_TRACKER!D$2:D$1000)')
    ws.cell(row=21, column=3).number_format = '$#,##0'
    ws.cell(row=21, column=4, value='=IFERROR(C21/B21,0)')
    ws.cell(row=21, column=4).number_format = '$#,##0'
    ws.cell(row=21, column=5, value='=IFERROR(SUMPRODUCT((REVENUE_TRACKER!P$2:P$1000="NON-AI")*(REVENUE_TRACKER!D$2:D$1000>=2500))/B21,0)')
    ws.cell(row=21, column=5).number_format = '0%'
    ws.cell(row=21, column=6, value='=IFERROR(MAXIFS(REVENUE_TRACKER!D$2:D$1000,REVENUE_TRACKER!P$2:P$1000,"NON-AI"),0)')
    ws.cell(row=21, column=6).number_format = '$#,##0'

    # === SECTION 4: INSTALL ECONOMICS (rows 24-28) ===
    ws['A24'] = 'INSTALL ECONOMICS'
    ws['A24'].font = section_font

    ws.cell(row=25, column=1, value='Avg revenue per install (all)')
    ws.cell(row=25, column=2, value='=IFERROR(SUM(REVENUE_TRACKER!D:D)/SUM(REVENUE_TRACKER!H:H),0)')
    ws.cell(row=25, column=2).number_format = '$#,##0.00'

    ws.cell(row=26, column=1, value='Avg revenue per install (earning only)')
    ws.cell(row=26, column=2, value='=IFERROR(SUMPRODUCT((REVENUE_TRACKER!D$2:D$1000>0)*REVENUE_TRACKER!D$2:D$1000)/SUMPRODUCT((REVENUE_TRACKER!D$2:D$1000>0)*REVENUE_TRACKER!H$2:H$1000),0)')
    ws.cell(row=26, column=2).number_format = '$#,##0.00'

    ws.cell(row=27, column=1, value='US/EU installs as % of total')
    ws.cell(row=27, column=2, value='=IFERROR(SUM(REVENUE_TRACKER!F:F)/SUM(REVENUE_TRACKER!H:H),0)')
    ws.cell(row=27, column=2).number_format = '0%'

    ws.cell(row=28, column=1, value='Min installs to hit cap')
    ws.cell(row=28, column=2, value='=IFERROR(MIN(IF(REVENUE_TRACKER!D$2:D$1000>=2500,REVENUE_TRACKER!H$2:H$1000)),0)')
    ws.cell(row=28, column=2).number_format = '#,##0'

    # Column widths
    for col, w in [('A', 35), ('B', 14), ('C', 16), ('D', 14), ('E', 12), ('F', 12)]:
        ws.column_dimensions[col].width = w


def _build_competitor_view_tab(ws, competitor_gaps, header_fill, header_font, thin_border):
    headers = ['Date', 'Competitor', 'Trend', 'Competitor Momentum', 'Your Momentum',
               'Competitor Shares/h', 'Market', 'Gap Type', 'Hours Behind',
               'Est. Missed Revenue ($)', 'AI Category', 'URL']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = header_fill; c.font = header_font; c.alignment = Alignment(horizontal='center')

    date_str = datetime.now().strftime('%Y-%m-%d')
    if len(competitor_gaps) > 0:
        for ri, (_, row) in enumerate(competitor_gaps.iterrows(), 2):
            vals = [date_str, row.get('competitor_account',''), str(row.get('trend_text',''))[:60],
                    int(row.get('competitor_momentum',0)), 0, round(float(row.get('competitor_shares_h',0)),1),
                    str(row.get('market','')), row.get('gap_type',''), row.get('hours_behind',''),
                    round(float(row.get('estimated_missed_revenue',0)),2), str(row.get('ai_category','')),
                    str(row.get('trend_url',''))]
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=_sanitize_cell(val))
                c.border = thin_border
            gap_cell = ws.cell(row=ri, column=8)
            if gap_cell.value == 'MISSED_BY_YOU':
                gap_cell.fill = PatternFill('solid', fgColor='FF6B6B')
                gap_cell.font = Font(bold=True, color='FFFFFF')
            elif gap_cell.value == 'BOTH_CAUGHT':
                gap_cell.fill = PatternFill('solid', fgColor='90EE90')
            url_cell = ws.cell(row=ri, column=12)
            uv = str(row.get('trend_url',''))
            if uv.startswith('http'):
                url_cell.hyperlink = uv
                url_cell.font = Font(color='0000FF', underline='single')
    else:
        ws['A2'] = 'No competitor posts found in today\'s trending data'
        ws['A2'].font = Font(italic=True, color='666666')
    ws.freeze_panes = 'A2'
    for col, w in [('A',12),('B',20),('C',50),('D',18),('E',15),('F',16),('G',15),('H',18),('I',14),('J',18),('K',12),('L',50)]:
        ws.column_dimensions[col].width = w


def _build_prediction_log_tab(ws, df_pred, df_yesterday, existing_log, header_fill, header_font, thin_border):
    headers = ['Date', 'Trends Tracked', 'Direction Accuracy %', 'Bias', 'MAPE %',
               'Correct Builds', 'False Positives', 'Missed Opportunities', 'Correct Skips', 'Tuning Suggestion']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = header_fill; c.font = header_font; c.alignment = Alignment(horizontal='center')

    start_row = 2
    alt_fill = PatternFill('solid', fgColor='F5F5F5')
    if existing_log is not None and len(existing_log) > 0:
        for ri, (_, row) in enumerate(existing_log.iterrows(), 2):
            for ci in range(1, 11):
                val = row.iloc[ci-1] if ci-1 < len(row) else ''
                if pd.isna(val): val = ''
                c = ws.cell(row=ri, column=ci, value=_sanitize_cell(val))
                c.border = thin_border
                if ri % 2 == 0: c.fill = alt_fill
            start_row = ri + 1

    date_str = datetime.now().strftime('%Y-%m-%d')
    if df_yesterday is not None and len(df_yesterday) > 0 and len(df_pred) > 0:
        tracked = len(df_pred[df_pred['prediction_confidence'].isin(['HIGH', 'MEDIUM'])])
        if tracked > 0:
            correct_dir = total_cmp = 0
            for _, row in df_pred.iterrows():
                vel = row.get('velocity', 0)
                delta = row.get('momentum_score', 0) - row.get('yesterday_momentum', 0)
                if pd.notna(vel) and pd.notna(row.get('yesterday_momentum')):
                    total_cmp += 1
                    if (vel > 0 and delta > 0) or (vel <= 0 and delta <= 0): correct_dir += 1
            dir_acc = correct_dir / total_cmp if total_cmp > 0 else 0
            vals = [date_str, tracked, round(dir_acc, 2), 'Neutral', 0, 0, 0, 0, 0, 'Insufficient data for tuning']
            new_row_fill = PatternFill('solid', fgColor='E0FFE0')  # Highlight today's new entry
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=start_row, column=ci, value=_sanitize_cell(val))
                c.border = thin_border
                c.fill = new_row_fill

    ws.freeze_panes = 'A2'
    for col, w in [('A',12),('B',15),('C',20),('D',15),('E',10),('J',40)]:
        ws.column_dimensions[col].width = w


def _build_data_feed_tab(ws, df_today, header_fill, header_font, thin_border, cyan_fill):
    headers = ['Date', 'Account', 'Trend', 'Age', 'Momentum', 'Status', 'Market',
               'Views/h', 'Shares/h', 'BUILD_NOW', 'TikTok URL', 'TUTORIAL_TRIGGER',
               'URGENCY', 'Trigger Reason', 'AI Category', 'Opportunity Score',
               'Time Zone', 'Build Priority', 'Seasonal Event']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = header_fill; c.font = header_font; c.alignment = Alignment(horizontal='center')

    date_str = datetime.now().strftime('%Y-%m-%d')
    if 'author' not in df_today.columns:
        ws['A2'] = 'No author data available'
        ws.freeze_panes = 'A2'
        return

    your_mask = df_today['author'].str.lower().isin([a.lower() for a in YOUR_ACCOUNTS])
    your_posts = df_today[your_mask].copy()
    now_hour = datetime.utcnow().hour
    tz_label = '\U0001f7e2 PRIME' if 8 <= now_hour <= 22 else 'OFF_PEAK'
    seasonal_text = ''
    try:
        from seasonal_calendar import get_seasonal_alerts
        from datetime import date
        alerts = get_seasonal_alerts(date.today())
        if alerts: seasonal_text = alerts[0].get('event', '')
    except Exception: pass

    from daily_processor import calculate_tutorial_trigger

    for ri, (_, row) in enumerate(your_posts.iterrows(), 2):
        trigger, urgency, reason = calculate_tutorial_trigger(row)
        mom = float(row.get('momentum_score', 0))
        shares_h = float(row.get('shares_per_hour', 0))
        age = float(row.get('age_hours', 0))
        opp_score = int(mom * 0.5 + shares_h * 10 + max(0, (72 - age)) * 5)
        if trigger == '\U0001f534 MAKE_NOW' and urgency == '\U0001f525 URGENT':
            build_pri = '\U0001f534 BUILD_IMMEDIATELY'
        elif trigger == '\U0001f534 MAKE_NOW':
            build_pri = '\U0001f7e0 BUILD_TODAY'
        elif trigger == '\U0001f7e1 WATCH':
            build_pri = '\U0001f7e1 MONITOR'
        else:
            build_pri = ''

        vals = [date_str, str(row.get('author','')), str(row.get('text',''))[:60] if pd.notna(row.get('text')) else '',
                f"{age:.1f}h", int(mom), str(row.get('acceleration_status', row.get('status',''))),
                str(row.get('Market','')) if pd.notna(row.get('Market')) else '',
                int(row.get('views_per_hour',0)), round(shares_h,1), str(row.get('BUILD_NOW','')),
                str(row.get('webVideoUrl','')), trigger, urgency, reason,
                str(row.get('AI_CATEGORY','')) if pd.notna(row.get('AI_CATEGORY')) else '',
                opp_score, tz_label, build_pri, seasonal_text]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=_sanitize_cell(val))
            c.border = thin_border
            if ci <= 11: c.fill = cyan_fill
        # Trigger colors
        trig_cell = ws.cell(row=ri, column=12)
        if 'MAKE_NOW' in str(trigger):
            trig_cell.fill = PatternFill('solid', fgColor='FF0000')
            trig_cell.font = Font(bold=True, color='FFFFFF')
        elif 'WATCH' in str(trigger):
            trig_cell.fill = PatternFill('solid', fgColor='FFFF00')
        urg_cell = ws.cell(row=ri, column=13)
        if 'URGENT' in str(urgency):
            urg_cell.fill = PatternFill('solid', fgColor='8B0000')
            urg_cell.font = Font(bold=True, color='FFFFFF')
        elif 'HIGH' in str(urgency):
            urg_cell.fill = PatternFill('solid', fgColor='FFA500')
        url_cell = ws.cell(row=ri, column=11)
        uv = str(row.get('webVideoUrl',''))
        if uv.startswith('http'):
            url_cell.hyperlink = uv
            url_cell.font = Font(color='0000FF', underline='single')

    ws.freeze_panes = 'A2'
    for col, w in [('A',12),('B',20),('C',50),('D',10),('E',12),('F',15),('G',15),('H',10),('I',10),('K',50),('L',18),('N',30)]:
        ws.column_dimensions[col].width = w


# =============================================================================
# REVENUE DATA PERSISTENCE
# =============================================================================

def _load_existing_revenue(dashboard_path, live_revenue_df=None):
    # Start with seed data as the base (always includes all known templates)
    seed_df = pd.DataFrame(SEED_REVENUE_DATA) if SEED_REVENUE_DATA else None
    
    # Determine the primary data source
    primary_df = None
    source = None
    
    # Priority 1: Live data from Google Sheet
    if live_revenue_df is not None and len(live_revenue_df) > 0:
        primary_df = live_revenue_df
        source = 'Google Sheet'
    
    # Priority 2: Existing dashboard file
    elif dashboard_path and os.path.exists(dashboard_path):
        try:
            wb = load_workbook(dashboard_path, data_only=True)
            if 'REVENUE_TRACKER' in wb.sheetnames:
                ws = wb['REVENUE_TRACKER']
                data = []
                headers = [ws.cell(1, c).value for c in range(1, 20)]
                for ri in range(2, ws.max_row + 1):
                    row_data = {}
                    has_data = False
                    for ci in range(1, 20):
                        val = ws.cell(ri, ci).value
                        col_name = headers[ci-1] if ci-1 < len(headers) else f'col_{ci}'
                        row_data[col_name] = val
                        if val is not None and str(val).strip() != '': has_data = True
                    if has_data: data.append(row_data)
                if data:
                    primary_df = pd.DataFrame(data)
                    source = 'dashboard file'
        except Exception as e:
            print(f"  Warning: Could not load revenue from file: {e}")
    
    # If no primary source, fall back to seed data alone
    if primary_df is None:
        if seed_df is not None:
            print(f"  Using embedded seed revenue data ({len(seed_df)} entries)")
            return seed_df
        return None
    
    # MERGE: primary data + any seed-only entries not in primary
    # This ensures templates tracked outside the Pioneer spreadsheet are never lost
    if seed_df is not None and len(seed_df) > 0:
        # Get URL column from primary
        primary_url_col = None
        for col in ['TikTok URL', 'url', 'tiktok_url', 'webVideoUrl']:
            if col in primary_df.columns:
                primary_url_col = col
                break
        
        seed_url_col = 'TikTok URL'  # Seed always uses this
        
        if primary_url_col and seed_url_col in seed_df.columns:
            primary_urls = set(primary_df[primary_url_col].astype(str).str.strip())
            
            # Find seed entries not in primary data
            seed_only = seed_df[~seed_df[seed_url_col].astype(str).str.strip().isin(primary_urls)].copy()
            
            if len(seed_only) > 0:
                # Rename seed columns to match primary if needed
                if primary_url_col != seed_url_col:
                    seed_only = seed_only.rename(columns={seed_url_col: primary_url_col})
                
                merged = pd.concat([primary_df, seed_only], ignore_index=True)
                seed_only_rev = seed_only['Received ($)'].sum() if 'Received ($)' in seed_only.columns else 0
                print(f"  Using live revenue data from {source} ({len(primary_df)} entries) + {len(seed_only)} seed-only entries (${seed_only_rev:,.0f} received)")
                return merged
    
    print(f"  Using live revenue data from {source} ({len(primary_df)} entries)")
    return primary_df


def _load_existing_prediction_log(dashboard_path):
    if not dashboard_path or not os.path.exists(dashboard_path):
        return None
    try:
        wb = load_workbook(dashboard_path, data_only=True)
        if 'PREDICTION_LOG' not in wb.sheetnames: return None
        ws = wb['PREDICTION_LOG']
        data = []
        for ri in range(2, ws.max_row + 1):
            row_data = [ws.cell(ri, c).value for c in range(1, 11)]
            if any(v is not None and str(v).strip() != '' for v in row_data):
                data.append(row_data)
        if not data: return None
        headers = [ws.cell(1, c).value for c in range(1, 11)]
        return pd.DataFrame(data, columns=headers)
    except Exception as e:
        print(f"  Warning: Could not load prediction log: {e}")
        return None


# =============================================================================
# DAILY BRIEFING - STRATEGIC ANALYSIS TEXT
# =============================================================================

def generate_daily_briefing(
    df_today: pd.DataFrame,
    df_yesterday: pd.DataFrame = None,
    output_dir: str = '.',
    cache_path: str = None
) -> str:
    """
    Generate written daily briefing with:
    1. Immediate Actions - top trends to build RIGHT NOW
    2. Strategic Insights - competitor analysis across ALL 6 competitors
    
    Returns the briefing text.
    """
    from datetime import datetime
    date_str = datetime.now().strftime('%Y-%m-%d')
    lines = []
    lines.append("=" * 60)
    lines.append(f"DAILY BRIEFING - {date_str}")
    lines.append("TikTok Trend System v3.6.0")
    lines.append("=" * 60)
    
    df = df_today.copy()
    df = _ensure_calculated_metrics(df)
    # Deduplicate by URL to prevent BOTH-market duplicates in briefing
    df = df.drop_duplicates(subset=['webVideoUrl'], keep='first')
    
    # --- SECTION 1: IMMEDIATE ACTIONS ---
    lines.append("")
    lines.append("━" * 60)
    lines.append("🔴 IMMEDIATE ACTIONS")
    lines.append("━" * 60)
    
    # Calculate velocity if yesterday available
    has_velocity = False
    if df_yesterday is not None and len(df_yesterday) > 0:
        velocity_df = calculate_velocity_predictions(df, df_yesterday)
        has_velocity = True
    else:
        velocity_df = df.copy()
        velocity_df['velocity'] = 0
        velocity_df['action_window'] = 'MONITOR'
        velocity_df['trajectory'] = 'FLAT'
        velocity_df['predicted_24h'] = velocity_df.get('momentum_score', 0)
    
    # Filter: fresh content under 48h with momentum >= 500
    fresh = velocity_df[velocity_df['age_hours'] <= 48].copy() if 'age_hours' in velocity_df.columns else velocity_df.copy()
    if 'momentum_score' in fresh.columns:
        fresh = fresh[fresh['momentum_score'] >= 500]
    
    # Exclude your own posts and competitor posts
    if 'author' in fresh.columns:
        all_tracked = [a.lower() for a in YOUR_ACCOUNTS + COMPETITOR_ACCOUNTS]
        fresh = fresh[~fresh['author'].str.lower().isin(all_tracked)]
    
    # Sort by momentum
    if len(fresh) > 0 and 'momentum_score' in fresh.columns:
        fresh = fresh.nlargest(5, 'momentum_score')
    
    if len(fresh) == 0:
        lines.append("")
        lines.append("No high-priority trends found meeting criteria (age <48h, momentum >=500).")
    else:
        for i, (_, row) in enumerate(fresh.iterrows(), 1):
            momentum = _safe_int_val(row.get('momentum_score', 0))
            age = round(float(row.get('age_hours', 0)), 1) if pd.notna(row.get('age_hours')) else 0
            shares_h = round(float(row.get('shares_per_hour', 0)), 1) if pd.notna(row.get('shares_per_hour')) else 0
            trend_text = str(row.get('text') if pd.notna(row.get('text')) else '')[:60]
            creator = str(row.get('author') if pd.notna(row.get('author')) else 'Unknown')[:20]
            market = str(row.get('Market', '')) if pd.notna(row.get('Market')) else ''
            
            # Action window
            action = str(row.get('action_window', '')) if has_velocity else ''
            trajectory = str(row.get('trajectory', '')) if has_velocity else ''
            vel = row.get('velocity', 0) if has_velocity else 0
            vel = vel if pd.notna(vel) else 0
            pred_24 = row.get('predicted_24h', momentum) if has_velocity else momentum
            pred_24 = pred_24 if pd.notna(pred_24) else momentum
            
            hours_left = max(0, 72 - age)
            
            lines.append("")
            lines.append(f"  #{i}. {trend_text}")
            lines.append(f"      Creator: {creator} | {market}")
            lines.append(f"      Momentum: {momentum:,} | Shares/h: {shares_h}")
            lines.append(f"      Age: {age}h | Window: {hours_left:.0f}h remaining")
            if has_velocity:
                lines.append(f"      Velocity: {vel:+,.0f}/day | Predicted 24h: {int(pred_24):,}")
                lines.append(f"      Action: {action} | Trajectory: {trajectory}")
            
            # Why this trend
            reasons = []
            if momentum >= 3000: reasons.append("URGENT momentum")
            elif momentum >= 2000: reasons.append("HIGH momentum")
            if shares_h >= 100: reasons.append(f"very high share rate ({shares_h}/h)")
            elif shares_h >= 25: reasons.append(f"strong share rate ({shares_h}/h)")
            if '🌐 BOTH' in market: reasons.append("trending in BOTH markets (2x revenue potential)")
            if has_velocity and vel > 200: reasons.append("EXPLOSIVE growth trajectory")
            elif has_velocity and vel > 100: reasons.append("strong upward velocity")
            if hours_left < 24: reasons.append(f"only {hours_left:.0f}h left in 72h window")
            
            if reasons:
                lines.append(f"      Why: {'; '.join(reasons)}")
    
    # --- SECTION 2: COMPETITOR ANALYSIS ---
    lines.append("")
    lines.append("━" * 60)
    lines.append("📊 STRATEGIC INSIGHTS - COMPETITOR ANALYSIS")
    lines.append("━" * 60)
    
    # Find all competitor posts
    comp_posts = pd.DataFrame()
    your_posts = pd.DataFrame()
    if 'author' in df.columns:
        comp_mask = df['author'].str.lower().isin([a.lower() for a in COMPETITOR_ACCOUNTS])
        your_mask = df['author'].str.lower().isin([a.lower() for a in YOUR_ACCOUNTS])
        comp_posts = df[comp_mask].copy()
        your_posts = df[your_mask].copy()
    
    lines.append("")
    lines.append(f"  Your posts in trending: {len(your_posts)}")
    lines.append(f"  Competitor posts in trending: {len(comp_posts)}")
    
    if len(comp_posts) == 0:
        lines.append("")
        lines.append("  No competitor posts found in today's trending data.")
        lines.append("  This could mean they're not posting, or their posts aren't trending.")
    else:
        # Per-competitor breakdown
        lines.append("")
        lines.append("  COMPETITOR BREAKDOWN:")
        comp_by_account = comp_posts.groupby('author').agg(
            posts=('author', 'size'),
            avg_momentum=('momentum_score', 'mean'),
            max_momentum=('momentum_score', 'max'),
            total_momentum=('momentum_score', 'sum')
        ).sort_values('total_momentum', ascending=False)
        
        for account, row in comp_by_account.iterrows():
            lines.append(f"    {account}: {int(row['posts'])} posts | "
                        f"avg momentum {int(row['avg_momentum']):,} | "
                        f"best {int(row['max_momentum']):,}")
        
        # Gap analysis - trends competitors caught that you missed
        lines.append("")
        lines.append("  GAP ANALYSIS - TRENDS THEY CAUGHT, YOU DIDN'T:")
        
        # Get your URLs
        your_urls = set(your_posts['webVideoUrl']) if len(your_posts) > 0 else set()
        
        # For each competitor post, find the underlying trend
        # A "missed" trend is where a competitor posted on a trending topic but none of your accounts did
        # We approximate this by checking if your accounts appear for similar content
        # Simpler: just show their highest momentum posts you don't have
        missed = comp_posts[~comp_posts['webVideoUrl'].isin(your_urls)].copy()
        missed = missed.nlargest(min(5, len(missed)), 'momentum_score')
        
        if len(missed) == 0:
            lines.append("    None! You covered all trends they did. 🎯")
        else:
            total_missed_revenue = 0
            for _, row in missed.iterrows():
                m = _safe_int_val(row.get('momentum_score', 0))
                text = str(row.get('text') if pd.notna(row.get('text')) else '')[:50]
                account = str(row.get('author', ''))
                age = round(float(row.get('age_hours', 0)), 1) if pd.notna(row.get('age_hours')) else 0
                from revenue_model import estimate_competitor_revenue as _est_comp2
                _est2 = _est_comp2(m, row.get('shares_per_hour'), row.get('age_hours'))
                est_rev = _est2['estimated_revenue']
                total_missed_revenue += est_rev
                
                lines.append(f"    • {text}")
                lines.append(f"      By: {account} | Momentum: {m:,} | Age: {age}h | Est. missed: £{est_rev:.0f}")
            
            lines.append(f"    Total estimated missed revenue: £{total_missed_revenue:.0f}")
        
        # Head-to-head comparison
        lines.append("")
        lines.append("  HEAD-TO-HEAD SCORECARD:")
        your_avg_m = int(your_posts['momentum_score'].mean()) if len(your_posts) > 0 else 0
        comp_avg_m = int(comp_posts['momentum_score'].mean()) if len(comp_posts) > 0 else 0
        your_total = int(your_posts['momentum_score'].sum()) if len(your_posts) > 0 else 0
        comp_total = int(comp_posts['momentum_score'].sum()) if len(comp_posts) > 0 else 0
        
        your_spiking = len(your_posts[your_posts.get('status', your_posts.get('acceleration_status', pd.Series())).str.contains('SPIKING', na=False)]) if len(your_posts) > 0 else 0
        comp_spiking = len(comp_posts[comp_posts.get('status', comp_posts.get('acceleration_status', pd.Series())).str.contains('SPIKING', na=False)]) if len(comp_posts) > 0 else 0
        
        lines.append(f"    Metric              YOU          COMPETITORS")
        lines.append(f"    Posts in trending    {len(your_posts):<12} {len(comp_posts)}")
        lines.append(f"    Avg momentum        {your_avg_m:<12,} {comp_avg_m:,}")
        lines.append(f"    Total momentum      {your_total:<12,} {comp_total:,}")
        lines.append(f"    SPIKING posts       {your_spiking:<12} {comp_spiking}")
        
        if your_total > comp_total:
            lines.append(f"    → You're WINNING overall ({your_total:,} vs {comp_total:,})")
        elif comp_total > your_total:
            lines.append(f"    → Competitors AHEAD ({comp_total:,} vs {your_total:,}) - find more trends!")
        else:
            lines.append(f"    → Even match")
        
        # Niche analysis
        if 'AI_CATEGORY' in comp_posts.columns:
            comp_ai = len(comp_posts[comp_posts['AI_CATEGORY'] == 'AI'])
            comp_non = len(comp_posts[comp_posts['AI_CATEGORY'] == 'NON-AI'])
            your_ai = len(your_posts[your_posts['AI_CATEGORY'] == 'AI']) if len(your_posts) > 0 and 'AI_CATEGORY' in your_posts.columns else 0
            your_non = len(your_posts[your_posts['AI_CATEGORY'] == 'NON-AI']) if len(your_posts) > 0 and 'AI_CATEGORY' in your_posts.columns else 0
            
            lines.append("")
            lines.append("  NICHE COVERAGE:")
            lines.append(f"    AI trends:     YOU {your_ai} vs COMP {comp_ai}")
            lines.append(f"    NON-AI trends: YOU {your_non} vs COMP {comp_non}")
            
            if comp_ai > your_ai * 2 and comp_ai >= 3:
                lines.append(f"    ⚠️ Competitors are covering more AI trends - consider increasing AI template output")
            if comp_non > your_non * 2 and comp_non >= 3:
                lines.append(f"    ⚠️ Competitors are covering more NON-AI trends - diversify beyond AI")
        
        # Market coverage
        if 'Market' in comp_posts.columns:
            comp_both = len(comp_posts[comp_posts['Market'].str.contains('BOTH', na=False)])
            your_both = len(your_posts[your_posts['Market'].str.contains('BOTH', na=False)]) if len(your_posts) > 0 and 'Market' in your_posts.columns else 0
            
            if comp_both > 0:
                lines.append("")
                lines.append(f"  CROSS-MARKET: Competitors have {comp_both} BOTH-market posts vs your {your_both}")
                if comp_both > your_both:
                    lines.append(f"    ⚠️ They're better at catching cross-market trends (2x revenue potential)")
    
    # --- SECTION 3: RECOMMENDATIONS ---
    lines.append("")
    lines.append("━" * 60)
    lines.append("💡 RECOMMENDATIONS")
    lines.append("━" * 60)
    lines.append("")
    
    recs = []
    if len(fresh) > 0:
        top = fresh.iloc[0]
        top_text = str(top.get('text') if pd.notna(top.get('text')) else 'Unknown trend')[:40]
        recs.append(f"1. BUILD NOW: Start with \"{top_text}\" - highest priority opportunity")
    
    if len(comp_posts) > 0 and len(missed) > 0:
        recs.append(f"2. Check {len(missed)} trends competitors caught that you missed - potential revenue gap")
    
    if len(your_posts) == 0:
        recs.append("3. ⚠️ None of your posts are in today's trending - check posting schedule")
    
    if has_velocity:
        explosive = velocity_df[velocity_df.get('trajectory', pd.Series()) == 'EXPLOSIVE'] if 'trajectory' in velocity_df.columns else pd.DataFrame()
        if len(explosive) > 0:
            recs.append(f"4. {len(explosive)} EXPLOSIVE trajectories detected - these will peak within 24h")
    
    if not recs:
        recs.append("Continue monitoring - no urgent action items today")
    
    for r in recs:
        lines.append(f"  {r}")
    
    lines.append("")
    lines.append("=" * 60)
    lines.append("END OF DAILY BRIEFING")
    lines.append("=" * 60)
    
    return "\n".join(lines)


def _safe_int_val(val, default=0):
    """Safe int conversion for briefing text."""
    try:
        if pd.isna(val):
            return default
        return int(val)
    except (ValueError, TypeError):
        return default


# =============================================================================
# INTEGRATION WITH EXISTING SYSTEM
# =============================================================================

def integrate_with_daily_processor(
    us_data: pd.DataFrame,
    uk_data: pd.DataFrame,
    yesterday_us: pd.DataFrame = None,
    yesterday_uk: pd.DataFrame = None,
    two_days_us: pd.DataFrame = None,
    two_days_uk: pd.DataFrame = None,
    output_dir: str = '.',
    dashboard_path: str = None,
    live_revenue_df: pd.DataFrame = None
) -> Dict[str, str]:
    """Main integration function - generates v3.6.0 7-tab enhanced files."""
    date_str = datetime.now().strftime('%Y-%m-%d')
    output_files = {}
    cache_dir = os.environ.get('CACHE_DIR', 'data')
    streak_cache_path = os.path.join(cache_dir, 'velocity_streak_cache.json')

    # Auto-detect dashboard file
    if not dashboard_path:
        for candidate in [
            os.path.join(cache_dir, 'TikTok_Dashboard_With_Revenue.xlsx'),
            os.path.join(output_dir, 'TikTok_Dashboard_With_Revenue.xlsx'),
            'data/TikTok_Dashboard_With_Revenue.xlsx',
        ]:
            if os.path.exists(candidate):
                dashboard_path = candidate
                print(f"  Found existing dashboard: {candidate}")
                break

    if us_data is not None and len(us_data) > 0:
        us_path = f"{output_dir}/BUILD_TODAY_US_ENHANCED_{date_str}.xlsx"
        create_enhanced_excel(us_data, yesterday_us, two_days_us, us_path,
                              cache_path=streak_cache_path, dashboard_path=dashboard_path,
                              live_revenue_df=live_revenue_df)
        output_files['us_enhanced'] = us_path

    if uk_data is not None and len(uk_data) > 0:
        uk_path = f"{output_dir}/BUILD_TODAY_UK_ENHANCED_{date_str}.xlsx"
        create_enhanced_excel(uk_data, yesterday_uk, two_days_uk, uk_path,
                              cache_path=streak_cache_path, dashboard_path=dashboard_path,
                              live_revenue_df=live_revenue_df)
        output_files['uk_enhanced'] = uk_path

    if us_data is not None and uk_data is not None:
        combined = pd.concat([us_data, uk_data], ignore_index=True)
        combined_yesterday = None
        if yesterday_us is not None and yesterday_uk is not None:
            combined_yesterday = pd.concat([yesterday_us, yesterday_uk], ignore_index=True)
            combined_yesterday = combined_yesterday.drop_duplicates(subset=['webVideoUrl'], keep='first')
        combined_path = f"{output_dir}/BUILD_TODAY_COMBINED_ENHANCED_{date_str}.xlsx"
        create_enhanced_excel(combined, combined_yesterday, None, combined_path,
                              cache_path=streak_cache_path, dashboard_path=dashboard_path,
                              live_revenue_df=live_revenue_df)
        output_files['combined_enhanced'] = combined_path

    return output_files


# =============================================================================
# STANDALONE TESTING
# =============================================================================

if __name__ == '__main__':
    # Test with sample data
    print("Creating sample test data...")
    
    # Generate sample data
    np.random.seed(42)
    n_samples = 100
    
    sample_data = pd.DataFrame({
        'webVideoUrl': [f'https://tiktok.com/video/{i}' for i in range(n_samples)],
        'text': [f'Sample trend #{i} with #capcut #trending hashtags' for i in range(n_samples)],
        'author': np.random.choice(
            YOUR_ACCOUNTS + COMPETITOR_ACCOUNTS + ['random_user_' + str(i) for i in range(20)],
            n_samples
        ),
        'age_hours': np.random.uniform(1, 72, n_samples),
        'shareCount': np.random.randint(10, 10000, n_samples),
        'diggCount': np.random.randint(100, 50000, n_samples),
        'playCount': np.random.randint(1000, 1000000, n_samples),
        'momentum_score': np.random.uniform(100, 5000, n_samples),
        'shares_per_hour': np.random.uniform(1, 200, n_samples),
        'views_per_hour': np.random.uniform(100, 50000, n_samples),
        'Market': np.random.choice(['🌐 BOTH', '🇺🇸 US ONLY', '🇬🇧 UK ONLY'], n_samples),
        'AI_CATEGORY': np.random.choice(['AI', 'NON-AI'], n_samples),
        'acceleration_status': np.random.choice(
            ['🆕 NEW', '🚀 SPIKING', '📈 RISING', '📉 COOLING', '❄️ DYING'],
            n_samples
        )
    })
    
    # Create yesterday's data (slightly different momentum)
    yesterday_data = sample_data.copy()
    yesterday_data['momentum_score'] = sample_data['momentum_score'] * np.random.uniform(0.7, 1.0, n_samples)
    
    # Run analysis
    print("Running velocity prediction analysis...")
    output_path = create_enhanced_excel(
        sample_data,
        yesterday_data,
        None,
        'test_enhanced_output.xlsx'
    )
    
    print(f"✅ Test file created: {output_path}")
    print("\nTab summary:")
    print("1. VELOCITY_PREDICTIONS - Where trends are heading")
    print("2. COMPETITOR_ANALYSIS - Gap analysis vs capcutdailyuk")
    print("3. HEAD_TO_HEAD - You vs Competitor scorecard")
    print("4. OPPORTUNITY_MATRIX - Your BUILD list for today")
    print("5. README - Documentation")
